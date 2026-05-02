# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import difflib
import importlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
from collections import defaultdict
from pathlib import Path

from dotenv import load_dotenv


REPO_ROOT = Path(__file__).resolve().parent
ENV_PATH = REPO_ROOT / ".env"
ENV_EXAMPLE_PATH = REPO_ROOT / ".env.example"
OUTPUT_DIR = REPO_ROOT / "output"
CATEGORY_BUCKET_PATH = REPO_ROOT / "lead_finder" / "category_bucket.json"
BUCKET_TEMPLATE_PATH = REPO_ROOT / "lead_finder" / "bucket_email_template.json"
STAT_PREFIX = "PIPELINE_STAT:"
EVENT_PREFIX = "PIPELINE_EVENT:"
ANSI_RESET = "\033[0m"
ANSI_BOLD = "\033[1m"
ANSI_DIM = "\033[2m"
ANSI_RED = "\033[31m"
ANSI_GREEN = "\033[32m"
ANSI_YELLOW = "\033[33m"
ANSI_BLUE = "\033[34m"
ANSI_MAGENTA = "\033[35m"
ANSI_CYAN = "\033[36m"
ANSI_GRAY = "\033[90m"
SUMMARY_METRICS = [
    ("Leads scraped", "leads_scraped", ANSI_CYAN),
    ("Sites built", "sites_built", ANSI_CYAN),
    ("Good", "sites_good", ANSI_GREEN),
    ("Rejected", "sites_rejected", ANSI_RED),
    ("WhatsApp YES", "whatsapp_yes", ANSI_YELLOW),
    ("Emails sent", "emails_sent", ANSI_MAGENTA),
]


def _enable_windows_ansi() -> None:
    if os.name != "nt":
        return
    try:
        import ctypes

        kernel32 = ctypes.windll.kernel32
        for handle_id in (-11, -12):
            handle = kernel32.GetStdHandle(handle_id)
            if handle in (0, -1):
                continue
            mode = ctypes.c_uint()
            if not kernel32.GetConsoleMode(handle, ctypes.byref(mode)):
                continue
            kernel32.SetConsoleMode(handle, mode.value | 0x0004)
    except Exception:
        pass


def _supports_color() -> bool:
    stream = sys.stdout
    if os.getenv("NO_COLOR"):
        return False
    if not hasattr(stream, "isatty") or not stream.isatty():
        return False
    if os.name == "nt":
        return True
    term = os.getenv("TERM", "")
    return term.lower() != "dumb"


_enable_windows_ansi()
COLOR_ENABLED = _supports_color()

load_dotenv(dotenv_path=ENV_PATH)


def _slugify(value: str) -> str:
    out = str(value or "").strip().lower()
    out = re.sub(r"[^a-z0-9]+", "-", out).strip("-")
    return out or "unknown"


def _style(text: str, *codes: str) -> str:
    if not COLOR_ENABLED or not codes:
        return text
    return f"{''.join(codes)}{text}{ANSI_RESET}"


def _status_label(status: str) -> str:
    normalized = status.upper()
    if normalized == "PASS":
        return _style(normalized, ANSI_BOLD, ANSI_GREEN)
    if normalized in {"FAIL", "ERROR"}:
        return _style(normalized, ANSI_BOLD, ANSI_RED)
    if normalized in {"WARN", "WARNING"}:
        return _style(normalized, ANSI_BOLD, ANSI_YELLOW)
    if normalized in {"INFO", "START"}:
        return _style(normalized, ANSI_BOLD, ANSI_CYAN)
    if normalized == "DONE":
        return _style(normalized, ANSI_BOLD, ANSI_GREEN)
    return _style(normalized, ANSI_BOLD, ANSI_BLUE)


def _status_color(status: str) -> str:
    normalized = status.upper()
    if normalized in {"SUCCESS", "DONE", "GOOD", "DEPLOYED", "SENT", "PASS"}:
        return ANSI_GREEN
    if normalized in {"FAIL", "FAILED", "REJECTED", "ERROR"}:
        return ANSI_RED
    if normalized in {"WARN", "WARNING", "PLACEHOLDERS_PENDING", "SKIPPED"}:
        return ANSI_YELLOW
    if normalized in {"START", "CHECK", "PLACEHOLDER_CHECK", "EXCEL_UPDATED"}:
        return ANSI_CYAN
    if normalized in {"PROGRESS", "TEMPLATE_FILL", "BUILT"}:
        return ANSI_BLUE
    return ANSI_MAGENTA if normalized == "EMAIL" else ANSI_BLUE


def _stage_color(stage: str) -> str:
    normalized = stage.upper()
    if normalized == "SCRAPE":
        return ANSI_CYAN
    if normalized == "BUILD":
        return ANSI_BLUE
    if normalized == "CHECK":
        return ANSI_GREEN
    if normalized == "SCREENSHOT":
        return ANSI_MAGENTA
    if normalized == "WHATSAPP":
        return ANSI_YELLOW
    if normalized == "EMAIL":
        return ANSI_MAGENTA
    if normalized == "SUMMARY":
        return ANSI_MAGENTA
    return ANSI_CYAN


def _print_banner(title: str, *, color: str = ANSI_CYAN) -> None:
    line = "=" * max(24, len(title) + 8)
    print()
    print(_style(line, ANSI_DIM, color))
    print(_style(f"=== {title} ===", ANSI_BOLD, color))
    print(_style(line, ANSI_DIM, color))


def _print_note(level: str, message: str) -> None:
    print(f"{_status_label(level)}: {message}")


def _print_city_header(city: str) -> None:
    print()
    print(_style(f"[CITY] {city}", ANSI_BOLD, ANSI_CYAN))


def _print_stage_header(stage: str) -> None:
    print(_style(f"[{stage.upper()}]", ANSI_BOLD, _stage_color(stage)))


def _parse_pipeline_event(line: str) -> dict[str, str] | None:
    raw = line.strip()
    if not raw.startswith(EVENT_PREFIX):
        return None
    payload = raw[len(EVENT_PREFIX) :].strip()
    if not payload:
        return None
    try:
        parsed = json.loads(payload)
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, dict) else None


def _print_pipeline_event(event: dict[str, str]) -> None:
    entity = str(event.get("entity") or "item").strip().upper()
    label = str(
        event.get("label")
        or event.get("id")
        or event.get("name")
        or event.get("path")
        or "-"
    ).strip() or "-"
    status = str(event.get("status") or "INFO").strip().upper()
    detail = str(event.get("detail") or "").strip()
    display = label[:40]
    dots = "." * max(4, 18 - len(display))
    suffix = f" ({detail})" if detail else ""
    print(
        f"  {_style(f'[{entity}]', ANSI_BOLD, _stage_color(str(event.get('stage') or '')))} "
        f"{display} {_style(dots, ANSI_DIM, ANSI_GRAY)} "
        f"{_style(status, ANSI_BOLD, _status_color(status))}{suffix}"
    )


def _child_env(base_env: dict[str, str] | None = None) -> dict[str, str]:
    merged = dict(os.environ)
    if base_env:
        merged.update(base_env)
    merged["PIPELINE_LOG_FORMAT"] = "structured"

    # Ensure REPO_ROOT is in PYTHONPATH so sub-scripts can find top-level modules like 'ai'
    pp = merged.get("PYTHONPATH", "")
    sep = os.pathsep
    new_pp = str(REPO_ROOT)
    if pp:
        new_pp = f"{new_pp}{sep}{pp}"
    merged["PYTHONPATH"] = new_pp

    return merged


def _should_hide_subprocess_line(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return True
    if stripped.startswith(STAT_PREFIX):
        return False

    lowered = stripped.lower()
    if re.match(r"^\d{4}-\d{2}-\d{2} ", stripped):
        if " warning " in f" {lowered} " or " error " in f" {lowered} ":
            return False
        return True
    if any(token in lowered for token in ("[warn]", "[error]", " traceback", "exception")):
        return False
    if re.search(r"\b(warn|warning|error|failed|fail)\b", lowered):
        return False

    noisy_prefixes = (
        "[category]",
        "[progress]",
        "[lead]",
        "[saved]",
        "[pause]",
        "[build]",
        "[skip]",
        "[push]",
        "[info]",
        "[ok]",
        "[dry-run]",
        "[dry-run-no-groq]",
        "[summary]",
        "[enrich]",
    )
    if lowered.startswith(noisy_prefixes):
        return True

    return False


def _style_subprocess_line(line: str) -> str:
    stripped = line.rstrip("\r\n")
    newline = line[len(stripped):]
    lowered = stripped.lower()
    if not stripped:
        return line

    if stripped.startswith(EVENT_PREFIX):
        return ""
    if stripped.startswith(STAT_PREFIX):
        return _style(stripped, ANSI_BOLD, ANSI_MAGENTA) + newline
    if "total:" in lowered:
        return _style(stripped, ANSI_BOLD, ANSI_BLUE) + newline
    if "[rejected]" in lowered or "[error]" in lowered or " traceback" in f" {lowered}" or "exception" in lowered:
        return _style(stripped, ANSI_BOLD, ANSI_RED) + newline
    if re.search(r"\b(fail|failed|error)\b", lowered):
        return _style(stripped, ANSI_BOLD, ANSI_RED) + newline
    if re.search(r"\b(warn|warning)\b", lowered):
        return _style(stripped, ANSI_BOLD, ANSI_YELLOW) + newline
    if "[good]" in lowered or "[ok]" in lowered or re.search(r"\bpass\b", lowered):
        return _style(stripped, ANSI_BOLD, ANSI_GREEN) + newline
    if re.search(r"\binfo\b", lowered) or "[info]" in lowered:
        return _style(stripped, ANSI_CYAN) + newline
    return line


def _cities_from_args(args: argparse.Namespace) -> list[str]:
    if args.city:
        return [args.city.strip()]
    return [item.strip() for item in str(args.cities or "").split(",") if item.strip()]


def _categories_from_file(path: Path) -> list[str]:
    lines = path.read_text(encoding="utf-8").splitlines()
    items: list[str] = []
    for raw in lines:
        line = raw.lstrip("\ufeff").strip()
        if line and not line.startswith("#"):
            items.append(line)
    return items


def _candidate_categories_paths(path: Path) -> list[Path]:
    raw = Path(path)
    candidates: list[Path] = []

    if raw.is_absolute():
        candidates.append(raw)
    else:
        candidates.append(REPO_ROOT / raw)
        candidates.append(REPO_ROOT / "lead_finder" / raw)
        if len(raw.parts) == 1:
            candidates.append(REPO_ROOT / "lead_finder" / raw.name)

    seen: set[Path] = set()
    unique: list[Path] = []
    for candidate in candidates:
        resolved = candidate.resolve(strict=False)
        if resolved in seen:
            continue
        seen.add(resolved)
        unique.append(candidate)
    return unique


def _find_similar_categories_file(path: Path) -> Path | None:
    lead_finder_dir = REPO_ROOT / "lead_finder"
    if not lead_finder_dir.is_dir():
        return None

    txt_files = sorted(item for item in lead_finder_dir.iterdir() if item.is_file() and item.suffix.lower() == ".txt")
    if not txt_files:
        return None

    names = [item.name for item in txt_files]
    matches = difflib.get_close_matches(path.name, names, n=1, cutoff=0.6)
    if not matches:
        return None

    matched_name = matches[0]
    return next((item for item in txt_files if item.name == matched_name), None)


def _resolve_categories_file(path: Path) -> tuple[Path | None, str]:
    candidates = _candidate_categories_paths(path)
    for candidate in candidates:
        if candidate.is_file():
            return candidate.resolve(), ""

    similar = _find_similar_categories_file(path)
    if similar is not None:
        return similar.resolve(), f"Using closest match: {similar.relative_to(REPO_ROOT)}"

    checked = ", ".join(str(candidate) for candidate in candidates)
    return None, checked


def _categories_from_args(args: argparse.Namespace) -> list[str]:
    if args.categories_file:
        return _categories_from_file(args.categories_file)
    if not args.categories:
        return []
    return [item.strip() for item in str(args.categories).split(",") if item.strip()]


def _validate_args(parser: argparse.ArgumentParser, args: argparse.Namespace) -> None:
    if args.max < 0:
        parser.error("--max must be 0 or greater.")

    if args.preflight:
        return

    if not _cities_from_args(args):
        parser.error("Use --city or --cities unless running with --preflight.")

    if not args.categories and not args.categories_file:
        parser.error("Use either --categories or --categories-file.")

    if args.categories_file:
        if args.categories_file.suffix.lower() != ".txt":
            parser.error("--categories-file must point to a .txt file.")
        resolved_categories_file, resolution_detail = _resolve_categories_file(args.categories_file)
        if resolved_categories_file is None:
            parser.error(
                "--categories-file was not found. "
                f"Received: {args.categories_file}. Checked: {resolution_detail}"
            )
        if resolution_detail:
            _print_note("INFO", resolution_detail)
        args.categories_file = resolved_categories_file

    try:
        categories = _categories_from_args(args)
    except OSError as exc:
        parser.error(f"Could not read --categories-file: {exc}")

    if not categories:
        parser.error("Provide at least one category via --categories or --categories-file.")


def _resolve_country_slug_for_city(city_input: str) -> str:
    city_slug = _slugify(city_input)
    cache_path = REPO_ROOT / "lead_finder" / "public" / "data" / "_city_country_cache.json"
    if cache_path.exists():
        try:
            raw = json.loads(cache_path.read_text(encoding="utf-8"))
            if isinstance(raw, dict) and city_slug in raw:
                return str(raw[city_slug] or "")
        except Exception:
            pass

    data_root = REPO_ROOT / "lead_finder" / "public" / "data"
    if not data_root.exists():
        return ""
    for entry in sorted(data_root.iterdir()):
        if not entry.is_dir():
            continue
        maybe_city = entry / city_slug
        if maybe_city.exists() and maybe_city.is_dir():
            return entry.name
    return ""


def _parse_pipeline_stat(line: str) -> tuple[str, int] | None:
    raw = line.strip()
    if not raw.startswith(STAT_PREFIX):
        return None
    payload = raw[len(STAT_PREFIX) :].strip()
    if "=" not in payload:
        return None
    key, value = payload.split("=", 1)
    try:
        return key.strip(), int(value.strip())
    except ValueError:
        return None


def _parse_excel_style_summary(line: str) -> dict[str, int]:
    stats: dict[str, int] = {}
    if "TOTAL:" not in line:
        return stats

    pairs = {
        key.strip().lower(): int(value)
        for key, value in re.findall(r"([A-Za-z_]+)=([0-9]+)", line)
    }

    if "good" in pairs:
        stats["sites_good"] = pairs["good"]
    if "rejected" in pairs:
        stats["sites_rejected"] = pairs["rejected"]
    if "yes" in pairs:
        stats["whatsapp_yes"] = pairs["yes"]
    return stats


def _parse_step_stats(line: str) -> dict[str, int]:
    stats: dict[str, int] = {}

    pipeline_stat = _parse_pipeline_stat(line)
    if pipeline_stat is not None:
        key, value = pipeline_stat
        stats[key] = value

    for key, value in _parse_excel_style_summary(line).items():
        stats[key] = value

    return stats


def _merge_stats(target: dict[str, int], incoming: dict[str, int]) -> None:
    for key, value in incoming.items():
        target[key] = target.get(key, 0) + value


def _stream_step_output(proc: subprocess.Popen, step_stats: dict[str, int]) -> None:
    assert proc.stdout is not None
    for line in proc.stdout:
        event = _parse_pipeline_event(line)
        if event is not None:
            _print_pipeline_event(event)
            continue
        if _should_hide_subprocess_line(line):
            continue
        rendered = _style_subprocess_line(line)
        if not rendered:
            continue
        sys.stdout.write(rendered)
        sys.stdout.flush()
        _merge_stats(step_stats, _parse_step_stats(line))


def _start_background_step(
    *,
    cmd: list[str],
    env: dict[str, str] | None = None,
) -> dict[str, object]:
    step_stats: dict[str, int] = defaultdict(int)
    proc = subprocess.Popen(
        cmd,
        cwd=str(REPO_ROOT),
        env=_child_env(env),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
    )
    thread = threading.Thread(target=_stream_step_output, args=(proc, step_stats), daemon=True)
    thread.start()
    return {
        "proc": proc,
        "thread": thread,
        "start": time.perf_counter(),
        "step_stats": step_stats,
    }


def _wait_background_step(name: str, handle: dict[str, object]) -> tuple[float, dict[str, int]]:
    proc = handle["proc"]
    thread = handle["thread"]
    start = handle["start"]
    step_stats = handle["step_stats"]

    assert isinstance(proc, subprocess.Popen)
    assert isinstance(thread, threading.Thread)
    assert isinstance(start, float)
    assert isinstance(step_stats, dict)

    proc.wait()
    thread.join()
    elapsed = time.perf_counter() - start
    if proc.returncode != 0:
        _print_note("FAIL", f"{name} exited with code {proc.returncode} after {elapsed:.2f}s")
        raise SystemExit(proc.returncode)
    _print_note("DONE", f"{name} completed in {elapsed:.2f}s")
    return elapsed, dict(step_stats)


def _stop_background_step(handle: dict[str, object]) -> None:
    proc = handle["proc"]
    thread = handle["thread"]

    assert isinstance(proc, subprocess.Popen)
    assert isinstance(thread, threading.Thread)

    if proc.poll() is None:
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait(timeout=5)
    thread.join(timeout=5)


def _run_step(
    *,
    name: str,
    stage: str,
    cmd: list[str],
    city: str | None = None,
    env: dict[str, str] | None = None,
) -> tuple[float, dict[str, int]]:
    if city:
        _print_city_header(city)
    _print_stage_header(stage)
    start = time.perf_counter()
    step_stats: dict[str, int] = defaultdict(int)
    proc = subprocess.Popen(
        cmd,
        cwd=str(REPO_ROOT),
        env=_child_env(env),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
    )
    _stream_step_output(proc, step_stats)

    proc.wait()
    elapsed = time.perf_counter() - start
    if proc.returncode != 0:
        _print_note("FAIL", f"{name} exited with code {proc.returncode} after {elapsed:.2f}s")
        raise SystemExit(proc.returncode)
    _print_note("DONE", f"{name} completed in {elapsed:.2f}s")
    return elapsed, dict(step_stats)


def _print_check(name: str, ok: bool, detail: str = "") -> bool:
    status = _status_label("PASS" if ok else "FAIL")
    suffix = f" - {detail}" if detail else ""
    print(f"{status}: {name}{suffix}")
    return ok


def _print_env_var_check(
    display_name: str,
    *,
    present: bool,
    source_name: str = "",
) -> bool:
    status = _status_label("PASS" if present else "FAIL")
    suffix = ""
    if present and source_name and source_name != display_name:
        suffix = f" (using {source_name})"
    elif not present:
        suffix = " (missing)"
    print(f"{display_name:<22} ... {status}{suffix}")
    return present


def _first_present_env(*names: str) -> tuple[bool, str]:
    for name in names:
        value = os.getenv(name, "").strip()
        if value:
            return True, name
    return False, ""


def _check_output_writable(path: Path) -> tuple[bool, str]:
    try:
        path.mkdir(parents=True, exist_ok=True)
        fd, tmp_path = tempfile.mkstemp(prefix=".pipeline_write_test_", dir=str(path))
        os.close(fd)
        os.remove(tmp_path)
        return True, str(path)
    except OSError as exc:
        return False, str(exc)


def _command_name(base: str) -> str:
    return f"{base}.cmd" if os.name == "nt" else base


def _check_website_builder_node_modules() -> tuple[bool, str]:
    node_modules_dir = REPO_ROOT / "website-builder" / "node_modules"
    if node_modules_dir.is_dir():
        return True, str(node_modules_dir)
    return False, (
        f"{node_modules_dir} not found. Run `cd website-builder && npm install` first."
    )


def _vercel_cli_candidates() -> list[tuple[list[str], str]]:
    candidates: list[tuple[list[str], str]] = []
    seen: set[tuple[str, ...]] = set()

    def add(cmd: list[str], label: str) -> None:
        key = tuple(cmd)
        if key in seen:
            return
        seen.add(key)
        candidates.append((cmd, label))

    local_vercel = REPO_ROOT / "website-builder" / "node_modules" / ".bin" / _command_name("vercel")
    if local_vercel.exists():
        add([str(local_vercel)], "website-builder/node_modules/.bin/vercel")

    path_vercel = shutil.which(_command_name("vercel")) or shutil.which("vercel")
    if path_vercel:
        add([path_vercel], f"PATH vercel ({path_vercel})")

    return candidates


def _check_vercel_cli() -> tuple[bool, str]:
    candidates = _vercel_cli_candidates()
    if not candidates:
        return False, (
            "No Vercel launcher found. Install with `cd website-builder && npm install`, "
            "`scripts\\setup-node.cmd`, or `npm i -g vercel`."
        )

    failures: list[str] = []
    for cmd, label in candidates:
        try:
            completed = subprocess.run(
                [*cmd, "--version"],
                cwd=str(REPO_ROOT),
                capture_output=True,
                text=True,
                timeout=120,
                check=False,
            )
        except Exception as exc:  # noqa: BLE001
            failures.append(f"{label}: {exc}")
            continue

        detail = (completed.stdout or completed.stderr or "").strip()
        detail = detail.splitlines()[0] if detail else ""
        if completed.returncode == 0:
            return True, f"{label}{f' - {detail}' if detail else ''}"
        failures.append(f"{label}: {detail or f'exit code {completed.returncode}'}")

    return False, " | ".join(failures)


def _check_python_module(module_name: str) -> tuple[bool, str]:
    try:
        importlib.import_module(module_name)
    except Exception as exc:  # noqa: BLE001
        return False, str(exc)
    return True, module_name


def _check_json_file(path: Path) -> tuple[bool, str]:
    try:
        json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        return False, str(exc)
    return True, str(path)


def _run_preflight() -> int:
    _print_banner("PREFLIGHT", color=ANSI_MAGENTA)

    results: list[bool] = []
    env_exists = ENV_PATH.exists()

    results.append(_print_check(".env file exists", env_exists, str(ENV_PATH)))
    if not env_exists and ENV_EXAMPLE_PATH.exists():
        _print_note("WARN", ".env not found. Copy .env.example to .env and fill in your values.")

    env_checks = [
        ("EMAIL_SMTP_HOST", ("EMAIL_SMTP_HOST", "SMTP_HOST")),
        ("EMAIL_SMTP_USER", ("EMAIL_SMTP_USER", "SMTP_USER")),
        ("EMAIL_HOST_PASSWORD", ("EMAIL_HOST_PASSWORD", "EMAIL_SMTP_PASS", "SMTP_PASS")),
    ]
    print()
    print(_style("Required env vars:", ANSI_BOLD, ANSI_BLUE))
    for display_name, names in env_checks:
        ok, source = _first_present_env(*names)
        results.append(
            _print_env_var_check(
                display_name,
                present=ok,
                source_name=source,
            )
        )

    node_path = shutil.which("node")
    results.append(_print_check("node is callable", bool(node_path), node_path or "node not found in PATH"))
    npm_path = shutil.which(_command_name("npm")) or shutil.which("npm")
    results.append(_print_check("npm is callable", bool(npm_path), npm_path or "npm not found in PATH"))

    node_modules_ok, node_modules_detail = _check_website_builder_node_modules()
    results.append(_print_check("website-builder/node_modules exists", node_modules_ok, node_modules_detail))

    for module_name in ("playwright", "openpyxl"):
        ok, detail = _check_python_module(module_name)
        results.append(_print_check(f"python dependency {module_name}", ok, detail))

    writable_ok, writable_detail = _check_output_writable(OUTPUT_DIR)
    results.append(_print_check("output/ directory is writable", writable_ok, writable_detail))

    results.append(
        _print_check(
            "lead_finder/category_bucket.json exists",
            CATEGORY_BUCKET_PATH.exists(),
            str(CATEGORY_BUCKET_PATH),
        )
    )

    template_ok = BUCKET_TEMPLATE_PATH.exists()
    template_detail = str(BUCKET_TEMPLATE_PATH)
    if template_ok:
        template_ok, template_detail = _check_json_file(BUCKET_TEMPLATE_PATH)
    results.append(
        _print_check(
            "lead_finder/bucket_email_template.json is valid JSON",
            template_ok,
            template_detail,
        )
    )

    all_ok = all(results)
    _print_banner("PREFLIGHT PASS" if all_ok else "PREFLIGHT FAIL", color=ANSI_GREEN if all_ok else ANSI_RED)
    return 0 if all_ok else 1


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Run BizSiteGen full pipeline.")
    city_group = p.add_mutually_exclusive_group(required=False)
    city_group.add_argument("--city", help='Single city, e.g. "bengaluru"')
    city_group.add_argument("--cities", help='Comma-separated cities, e.g. "bengaluru,mysuru"')

    category_group = p.add_mutually_exclusive_group(required=False)
    category_group.add_argument("--categories", help='CSV categories, e.g. "salon,gym"')
    category_group.add_argument(
        "--categories-file",
        type=Path,
        help="Path to a .txt file with one category per non-empty line.",
    )

    p.add_argument("--max", type=int, default=0, help="Max results per category (0 = unlimited)")
    p.add_argument("--limit", type=int, default=None, help="Limit website-builder to N leads")
    p.add_argument("--id", default=None, help="Build only one shop_id in website-builder")
    p.add_argument("--batch", type=int, default=1, help="Website-builder batch size")
    p.add_argument("--dry-run", action="store_true", help="Skip deploy, and do dry-runs where supported")
    p.add_argument("--preview", action="store_true", help="Enable website-builder preview mode")
    p.add_argument("--analyze-websites", action="store_true", help="Enable lead_finder website analysis")
    p.add_argument("--show-browser", action="store_true", help="Show the scraper browser UI")
    p.add_argument(
        "--force",
        action="store_true",
        help="Force lead_finder to ignore scraped-place deduplication and collect fresh listings.",
    )
    p.add_argument(
        "--city-rate-limit-seconds",
        type=float,
        default=0.0,
        help="Pause between cities to avoid hammering external services.",
    )
    p.add_argument("--skip-scrape", action="store_true", help="Skip lead_finder step")
    p.add_argument("--skip-build", action="store_true", help="Skip website-builder step")
    p.add_argument("--skip-whatsapp", action="store_true", help="Skip WhatsApp checker step")
    p.add_argument("--skip-email", action="store_true", help="Skip email sender step")
    p.add_argument(
        "--pause-after-build",
        action="store_true",
        help="Stop after website-builder so operators can run review before outreach.",
    )
    p.add_argument(
        "--skip-review-gate",
        action="store_true",
        help="DANGEROUS: Skip the requirement for approved reviews before email_sender runs.",
    )
    p.add_argument(
        "--preflight",
        action="store_true",
        help="Check repo configuration and dependencies without running the pipeline.",
    )
    return p


def _build_lead_finder_cmd(args: argparse.Namespace, city: str, categories: list[str]) -> list[str]:
    cmd = [
        sys.executable,
        str(REPO_ROOT / "lead_finder" / "run.py"),
        "--city",
        city,
        "--categories",
        ",".join(categories),
        "--max",
        str(args.max),
    ]
    if args.analyze_websites:
        cmd.append("--analyze-websites")
    if args.show_browser:
        cmd.append("--show-browser")
    if args.force:
        cmd.append("--force")
    return cmd


def _build_website_builder_cmd(
    node: str,
    args: argparse.Namespace,
    city: str,
    categories: list[str],
) -> list[str]:
    cmd = [
        node,
        str(REPO_ROOT / "website-builder" / "src" / "index.js"),
        "run",  # Fixed: watch command doesn't exist, use run
        "--batch",
        str(args.batch),
    ]
    if args.limit is not None:
        cmd.extend(["--limit", str(args.limit)])
    if args.id:
        cmd.extend(["--id", str(args.id)])
    if args.dry_run:
        cmd.append("--dry-run")
    if args.preview:
        cmd.append("--preview")
    return cmd


def _build_website_builder_env(
    city: str,
    categories: list[str],
) -> dict[str, str]:
    env = dict(os.environ)
    env["USE_JSON_LEADS"] = "true"
    env["JSON_LEADS_CITY_SLUG"] = _slugify(city)

    allowed = [_slugify(category) for category in categories if category.strip()]
    if allowed:
        env["ANALYTICS_CATEGORY_FILTER"] = ",".join(allowed)
        env["JSON_LEADS_CATEGORY_FILTER"] = ",".join(allowed)

    country_slug = _resolve_country_slug_for_city(city)
    if country_slug:
        env["ANALYTICS_KEY_PREFIX"] = f"{country_slug}/{_slugify(city)}/"
        env["JSON_LEADS_COUNTRY_SLUG"] = country_slug
    return env


def _build_website_checker_cmd(args: argparse.Namespace, city_dir: Path | None = None) -> list[str]:
    cmd = [sys.executable, str(REPO_ROOT / "website_checker" / "run.py")]
    if city_dir is not None:
        cmd.extend(["--output-dir", str(city_dir)])
    if args.dry_run:
        cmd.append("--dry-run")
    return cmd


def _find_output_city_dirs(city: str) -> list[Path]:
    city_slug = _slugify(city)
    matches: list[Path] = []
    seen: set[Path] = set()

    preferred_country = _resolve_country_slug_for_city(city)
    preferred = OUTPUT_DIR / preferred_country / city_slug if preferred_country else None
    candidates = [preferred] if preferred is not None else []

    if OUTPUT_DIR.exists():
        for country_dir in sorted(OUTPUT_DIR.iterdir()):
            if not country_dir.is_dir():
                continue
            candidates.append(country_dir / city_slug)

    for candidate in candidates:
        if candidate is None or not candidate.is_dir():
            continue
        resolved = candidate.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        matches.append(candidate)

    return matches


def _build_whatsapp_cmd(args: argparse.Namespace, city_dir: Path | None = None) -> list[str]:
    cmd = [sys.executable, str(REPO_ROOT / "whatsappcheck" / "run.py")]
    if args.dry_run:
        cmd.append("--dry-run")
    if city_dir is not None:
        cmd.extend(["--output-dir", str(city_dir)])
    return cmd


def _build_screenshot_cmd(args: argparse.Namespace, city_dir: Path | None = None) -> list[str]:
    cmd = [sys.executable, str(REPO_ROOT / "screenshot-taker" / "run.py")]
    if args.dry_run:
        cmd.append("--dry-run")
    if city_dir is not None:
        cmd.extend(["--output-dir", str(city_dir)])
    return cmd


def _build_email_cmd(args: argparse.Namespace, city: str | None = None) -> list[str]:
    cmd = [sys.executable, str(REPO_ROOT / "email_sender" / "agent.py")]
    if city:
        cmd.append(city)
    if not args.skip_review_gate:
        cmd.append("--require-approved-review")
    if args.dry_run:
        cmd.extend(["--dry-run", "--dry-run-no-gen"])
    return cmd


def _read_output_summary_for_cities(cities: list[str]) -> dict[str, int]:
    try:
        import openpyxl
    except ImportError:
        return {
            "sites_good": 0,
            "sites_rejected": 0,
            "whatsapp_yes": 0,
        }

    totals = {
        "sites_good": 0,
        "sites_rejected": 0,
        "whatsapp_yes": 0,
    }
    seen_files: set[Path] = set()

    for city in cities:
        for city_dir in _find_output_city_dirs(city):
            for file_path in sorted(city_dir.rglob("leads.xlsx")):
                resolved = file_path.resolve()
                if resolved in seen_files:
                    continue
                seen_files.add(resolved)

                try:
                    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    worksheet = workbook.active
                    headers = {
                        str(worksheet.cell(1, column).value or "").strip().lower(): column
                        for column in range(1, worksheet.max_column + 1)
                    }
                    review_col = headers.get("review_status")
                    whatsapp_col = headers.get("whatsapp")

                    for row_idx in range(2, worksheet.max_row + 1):
                        if review_col is not None:
                            review_status = str(
                                worksheet.cell(row_idx, review_col).value or ""
                            ).strip().lower()
                            if review_status in {"good", "approved"}:
                                totals["sites_good"] += 1
                            elif review_status == "rejected":
                                totals["sites_rejected"] += 1

                        if whatsapp_col is not None:
                            whatsapp = str(
                                worksheet.cell(row_idx, whatsapp_col).value or ""
                            ).strip().upper()
                            if whatsapp == "YES":
                                totals["whatsapp_yes"] += 1
                except Exception as exc:  # noqa: BLE001
                    print(f"[warn] Could not read summary data from {file_path}: {exc}")
                finally:
                    try:
                        workbook.close()
                    except Exception:  # noqa: BLE001
                        pass

    return totals


def _print_summary(
    overall: float,
    cities: list[str],
    city_stats: dict[str, dict[str, int]],
    total_stats: dict[str, int],
) -> None:
    print()
    print(_style("[SUMMARY]", ANSI_BOLD, ANSI_MAGENTA))
    for city in cities:
        stats = city_stats.get(city, {})
        print(_style(f"  [CITY] {city}", ANSI_BOLD, ANSI_CYAN))
        for label, key, color in SUMMARY_METRICS:
            print(f"  {label}: {_style(str(stats.get(key, 0)), ANSI_BOLD, color)}")
        print()

    print(_style("  [TOTAL]", ANSI_BOLD, ANSI_MAGENTA))
    for label, key, color in SUMMARY_METRICS:
        print(f"  {label}: {_style(str(total_stats.get(key, 0)), ANSI_BOLD, color)}")
    print(f"  Duration: {_style(f'{overall:.2f}s', ANSI_BOLD, ANSI_BLUE)}")


def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()
    _validate_args(parser, args)

    if args.preflight:
        raise SystemExit(_run_preflight())

    cities = _cities_from_args(args)
    if not cities:
        raise SystemExit("No cities provided. Use --city/--cities or run with --preflight.")

    categories = _categories_from_args(args)

    node = None
    if not args.skip_build:
        node = shutil.which("node")
        if not node:
            raise SystemExit("node not found in PATH (required for website-builder step).")
        npm_path = shutil.which(_command_name("npm")) or shutil.which("npm")
        if not npm_path:
            raise SystemExit("npm not found in PATH (required for website-builder build step).")

        node_modules_ok, node_modules_detail = _check_website_builder_node_modules()
        if not node_modules_ok:
            raise SystemExit(node_modules_detail)

    timings: dict[str, float] = {}
    pipeline_stats: dict[str, int] = defaultdict(int)
    city_stats: dict[str, dict[str, int]] = {city: defaultdict(int) for city in cities}
    overall_start = time.perf_counter()

    for index, city in enumerate(cities, start=1):
        if not args.skip_scrape:
            step_name = f"Lead Finder (scrape) [{city}]"
            elapsed, step_stats = _run_step(
                name=step_name,
                stage="SCRAPE",
                cmd=_build_lead_finder_cmd(args, city, categories),
                city=city,
            )
            timings[step_name] = elapsed
            _merge_stats(pipeline_stats, step_stats)
            _merge_stats(city_stats[city], step_stats)

        if not args.skip_build:
            step_name = f"Website Builder (build) [{city}]"
            if args.skip_scrape:
                _print_city_header(city)
                _print_stage_header("BUILD")
            elapsed, step_stats = _run_step(
                name=step_name,
                stage="BUILD",
                cmd=_build_website_builder_cmd(
                    node=node or "node",
                    args=args,
                    city=city,
                    categories=categories,
                ),
                env=_build_website_builder_env(
                    city,
                    categories,
                ),
                city=city,
            )
            timings[step_name] = elapsed
            _merge_stats(pipeline_stats, step_stats)
            _merge_stats(city_stats[city], step_stats)

        should_wait = index < len(cities) and args.city_rate_limit_seconds > 0
        if should_wait:
            _print_note("INFO", f"Waiting {args.city_rate_limit_seconds:.1f}s before next city")
            time.sleep(args.city_rate_limit_seconds)

    if args.pause_after_build:
        overall = time.perf_counter() - overall_start
        _print_summary(overall, cities, city_stats, pipeline_stats)
        review_cmd = "node website-builder/src/index.js review"
        _print_banner("PAUSED AFTER BUILD", color=ANSI_YELLOW)
        _print_note("INFO", "Website generation is complete. Review the generated sites before outreach.")
        _print_note("INFO", f"Run: {review_cmd}")
        _print_note("INFO", "WhatsApp and email steps were skipped because --pause-after-build was set.")
        return

    for city in cities:
        city_dirs = _find_output_city_dirs(city)
        if not city_dirs:
            _print_note("WARN", f"No output city directories found for checker stage in {city}.")
            continue

        for city_dir in city_dirs:
            step_name = f"Website Checker [{city_dir.relative_to(OUTPUT_DIR)}]"
            elapsed, step_stats = _run_step(
                name=step_name,
                stage="CHECK",
                cmd=_build_website_checker_cmd(args, city_dir=city_dir),
                city=city,
            )
            timings[step_name] = elapsed
            _merge_stats(pipeline_stats, step_stats)
            _merge_stats(city_stats[city], step_stats)

    for city in cities:
        city_dirs = _find_output_city_dirs(city)
        if not city_dirs:
            _print_note("WARN", f"No output city directories found for screenshot stage in {city}.")
            continue

        for city_dir in city_dirs:
            step_name = f"Screenshot Taker [{city_dir.relative_to(OUTPUT_DIR)}]"
            elapsed, step_stats = _run_step(
                name=step_name,
                stage="SCREENSHOT",
                cmd=_build_screenshot_cmd(args, city_dir=city_dir),
                city=city,
            )
            timings[step_name] = elapsed
            _merge_stats(pipeline_stats, step_stats)
            _merge_stats(city_stats[city], step_stats)

    if not args.skip_whatsapp:
        for city in cities:
            city_dirs = _find_output_city_dirs(city)
            if not city_dirs:
                step_name = f"WhatsApp Checker [{city}]"
                elapsed, step_stats = _run_step(
                    name=step_name,
                    stage="WHATSAPP",
                    cmd=_build_whatsapp_cmd(args),
                    city=city,
                )
                timings[step_name] = elapsed
                _merge_stats(pipeline_stats, step_stats)
                _merge_stats(city_stats[city], step_stats)
                continue

            for city_dir in city_dirs:
                step_name = f"WhatsApp Checker [{city_dir.relative_to(OUTPUT_DIR)}]"
                elapsed, step_stats = _run_step(
                    name=step_name,
                    stage="WHATSAPP",
                    cmd=_build_whatsapp_cmd(args, city_dir=city_dir),
                    city=city,
                )
                timings[step_name] = elapsed
                _merge_stats(pipeline_stats, step_stats)
                _merge_stats(city_stats[city], step_stats)

    if not args.skip_email:
        for city in cities:
            step_name = f"Email Sender [{city}]"
            elapsed, step_stats = _run_step(
                name=step_name,
                stage="EMAIL",
                cmd=_build_email_cmd(args, city),
                city=city,
            )
            timings[step_name] = elapsed
            _merge_stats(pipeline_stats, step_stats)
            _merge_stats(city_stats[city], step_stats)

    overall = time.perf_counter() - overall_start
    _print_summary(overall, cities, city_stats, pipeline_stats)


if __name__ == "__main__":
    main()
