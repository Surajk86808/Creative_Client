# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import hashlib
import io
import json
import logging
import os
import re
import smtplib
import time
from datetime import datetime, timedelta
from email.message import EmailMessage
from email.utils import parseaddr
from email.utils import make_msgid
from pathlib import Path
from pathlib import Path as _Path
from typing import Any

import openpyxl
import requests
from dotenv import load_dotenv
import sys

try:
    from .audit_store import append_audit_log, append_jsonl, utc_now_iso
    from .guardrails import validate_generated_email
    from .rate_limiter import RateLimiter
    from .retry_utils import retry_operation
    from .validation import TemplateValidationError, validate_template_files
except ImportError:
    from audit_store import append_audit_log, append_jsonl, utc_now_iso
    from guardrails import validate_generated_email
    from rate_limiter import RateLimiter
    from retry_utils import retry_operation
    from validation import TemplateValidationError, validate_template_files

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


REPO_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(dotenv_path=REPO_ROOT / ".env")

sys.path.insert(0, str(REPO_ROOT))
from ai.model_config import MODEL_CONFIG

OUTPUT_DIR = _Path(os.getenv("OUTPUT_DIR", str(REPO_ROOT / "output"))).resolve()
LEAD_FINDER_PUBLIC_DIR = REPO_ROOT / "lead_finder" / "public"
EMAIL_PUBLIC_DIR = REPO_ROOT / "public"

NVIDIA_API_BASE = os.getenv("NVIDIA_API_BASE", "https://integrate.api.nvidia.com/v1")

# Tasks can use distinct keys if provided, otherwise fallback to the shared NVIDIA_API_KEY
NVIDIA_API_KEY = os.getenv("NVIDIA_API_KEY", "").strip()
NVIDIA_EMAIL_API_KEY = os.getenv("NVIDIA_EMAIL_API_KEY", NVIDIA_API_KEY).strip()
NVIDIA_GUARD_API_KEY = os.getenv("NVIDIA_GUARD_API_KEY", NVIDIA_API_KEY).strip()

SMTP_HOST = os.getenv("EMAIL_SMTP_HOST", "smtp.zoho.in")
SMTP_PORT = int(os.getenv("EMAIL_SMTP_PORT", "465"))
SMTP_SECURITY = os.getenv("EMAIL_SMTP_SECURITY", "ssl").strip().lower()
SMTP_USERNAME = os.getenv("EMAIL_SMTP_USER", "").strip()
SMTP_PASSWORD = (os.getenv("EMAIL_HOST_PASSWORD") or os.getenv("EMAIL_SMTP_PASS") or "").strip()
AGENCY_NAME = os.getenv("AGENCY_NAME", "").strip()
AGENCY_WEBSITE = os.getenv("AGENCY_WEBSITE", "").strip()
SENDER_PHONE = os.getenv("SENDER_PHONE", "").strip()
SENDER_NAME = os.getenv("SENDER_NAME", "").strip()
SIGNATURE = os.getenv("EMAIL_SIGNATURE", "").strip()
UNSUBSCRIBE_TEXT = os.getenv("EMAIL_UNSUBSCRIBE_TEXT", "To opt out, reply STOP.").strip()
_suppression_raw = os.getenv(
    "EMAIL_SUPPRESSION_LIST_PATH",
    "public/email_status/suppression_list.txt",
)
_suppression_path = Path(_suppression_raw)
SUPPRESSION_LIST_PATH = (
    _suppression_path
    if _suppression_path.is_absolute()
    else (REPO_ROOT / _suppression_path)
)
BLOCKED_EMAIL_DOMAINS = {
    item.strip().lower()
    for item in os.getenv("EMAIL_BLOCKED_DOMAINS", "").split(",")
    if item.strip()
}
MAX_PER_HOUR = int(os.getenv("MAX_PER_HOUR", "50"))
MAX_PER_DAY = int(os.getenv("MAX_PER_DAY", "200"))
MAX_FAILED_ATTEMPTS_PER_LEAD = int(os.getenv("MAX_FAILED_ATTEMPTS_PER_LEAD", "3"))
FAILED_RETRY_COOLDOWN_HOURS = int(os.getenv("FAILED_RETRY_COOLDOWN_HOURS", "24"))

# Cross-folder dependencies:
# - lead_finder/category_bucket.json maps business categories to email buckets/scenarios.
# - lead_finder/bucket_email_template.json stores the bucket/scenario email templates.
# This sender lives in email_sender/, but it depends on those shared lead_finder files.
CATEGORY_BUCKET_PATH = REPO_ROOT / "lead_finder" / "category_bucket.json"
BUCKET_TEMPLATE_PATH = REPO_ROOT / "lead_finder" / "bucket_email_template.json"


logger = logging.getLogger("email_sender")
ROOT_DIR = Path(__file__).resolve().parents[1]
DAILY_SUMMARY_PATH = ROOT_DIR / "email_send_summary.json"
EMAIL_REGEX = re.compile(
    r"^[A-Za-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[A-Za-z0-9!#$%&'*+/=?^_`{|}~-]+)*"
    r"@[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?"
    r"(?:\.[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?)+$"
)
PIPELINE_LOG_STRUCTURED = os.getenv("PIPELINE_LOG_FORMAT", "").strip().lower() == "structured"


def _emit_pipeline_event(entity: str, label: str, status: str, detail: str = "") -> bool:
    if not PIPELINE_LOG_STRUCTURED:
        return False
    payload = {
        "stage": "email",
        "entity": entity,
        "label": label,
        "status": status,
    }
    if detail:
        payload["detail"] = detail
    try:
        print(f"PIPELINE_EVENT: {json.dumps(payload, ensure_ascii=False)}", flush=True)
    except UnicodeEncodeError:
        safe_payload = {
            k: v.encode("ascii", "replace").decode() if isinstance(v, str) else v
            for k, v in payload.items()
        }
        print(f"PIPELINE_EVENT: {json.dumps(safe_payload)}", flush=True)
    return True


def _lead_event_label(lead: dict[str, Any]) -> str:
    for key in ("shop_name", "name", "business_name", "lead_id"):
        value = lead.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return "unknown"


def _sanitize_city_name(city_name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", city_name.strip().lower()).strip("-")
    if not slug:
        raise ValueError("City name must contain alphanumeric characters.")
    return slug


def _candidate_public_data_roots() -> list[Path]:
    roots = [
        EMAIL_PUBLIC_DIR / "data",
        LEAD_FINDER_PUBLIC_DIR / "data",
    ]
    unique: list[Path] = []
    seen: set[Path] = set()
    for root in roots:
        resolved = root.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        unique.append(root)
    return unique


def _email_status_root() -> Path:
    return EMAIL_PUBLIC_DIR / "email_status"


def _ensure_template_dependencies_exist() -> None:
    missing = [path for path in (CATEGORY_BUCKET_PATH, BUCKET_TEMPLATE_PATH) if not path.exists()]
    if not missing:
        return
    missing_paths = ", ".join(str(path) for path in missing)
    raise SystemExit(
        "Missing required shared email template file(s): "
        f"{missing_paths}. These files live under lead_finder/ and must exist "
        "before running email_sender/agent.py."
    )


def _count_approved_review_rows(output_root: Path) -> int:
    approved_rows = 0
    for file_path in sorted(output_root.rglob("leads.xlsx")):
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            worksheet = workbook.active
            headers = {
                str(worksheet.cell(1, column).value or "").strip().lower(): column
                for column in range(1, worksheet.max_column + 1)
            }
            review_col = headers.get("review_status")
            if review_col is None:
                workbook.close()
                continue
            for row_idx in range(2, worksheet.max_row + 1):
                value = str(worksheet.cell(row_idx, review_col).value or "").strip().lower()
                if value in {"approved", "good"}:
                    approved_rows += 1
            workbook.close()
        except Exception as exc:  # noqa: BLE001
            logger.warning("Failed to inspect review status in %s: %s", file_path, exc)
    return approved_rows


def _is_valid_email(email: str) -> bool:
    _, parsed = parseaddr(email)
    if not parsed or parsed != email:
        return False
    if ".." in parsed:
        return False
    return EMAIL_REGEX.match(parsed) is not None


def _parse_iso_timestamp(value: str) -> datetime | None:
    if not value:
        return None
    normalized = value.strip()
    if normalized.endswith("Z"):
        normalized = f"{normalized[:-1]}+00:00"
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        return None


def _build_signature() -> str:
    if SIGNATURE:
        return SIGNATURE
    if AGENCY_WEBSITE and SENDER_PHONE and SENDER_NAME:
        return f"Website: {AGENCY_WEBSITE}\nMobile: {SENDER_PHONE}\n{SENDER_NAME}"
    return ""


def _configure_logging() -> None:
    if not logging.getLogger().handlers:
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s %(levelname)s %(message)s",
        )


def _validate_required_config(*, signature: str, needs_generation: bool, needs_send: bool) -> None:
    required_values: dict[str, str] = {}
    if needs_generation:
        required_values.update(
            {
                "AGENCY_NAME": AGENCY_NAME,
                "AGENCY_WEBSITE": AGENCY_WEBSITE,
                "SENDER_PHONE": SENDER_PHONE,
                "SENDER_NAME": SENDER_NAME,
                "SIGNATURE or EMAIL_SIGNATURE": signature,
            }
        )
    if needs_send:
        required_values.update(
            {
                "EMAIL_SMTP_USER": SMTP_USERNAME,
                "EMAIL_HOST_PASSWORD": SMTP_PASSWORD,
            }
        )
    missing = [key for key, value in required_values.items() if not value]
    if missing:
        raise SystemExit(f"Missing required configuration: {', '.join(missing)}")


def _load_suppression_list(path: Path) -> set[str]:
    if not path.exists():
        return set()
    entries: set[str] = set()
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip().lower()
        if not line or line.startswith("#"):
            continue
        entries.add(line)
    return entries


def _email_domain(email: str) -> str:
    return email.rsplit("@", 1)[1].strip().lower() if "@" in email else ""


def _is_suppressed_recipient(email: str, suppression_list: set[str]) -> tuple[bool, str]:
    domain = _email_domain(email)
    candidates = [email]
    if domain:
        candidates.extend([domain, f"@{domain}"])
    for candidate in candidates:
        if candidate in suppression_list:
            if candidate == email:
                return True, "recipient in suppression list"
            return True, f"domain {domain} in suppression list"
    if domain and domain in BLOCKED_EMAIL_DOMAINS:
        return True, f"domain {domain} blocked"
    return False, ""


def _should_skip_due_to_failures(existing_status: dict[str, Any]) -> tuple[bool, str]:
    if not isinstance(existing_status, dict):
        return False, ""
    if existing_status.get("status") != "FAILED":
        return False, ""

    failed_attempts = _to_int_or_none(existing_status.get("failed_attempts")) or 0
    if failed_attempts >= MAX_FAILED_ATTEMPTS_PER_LEAD:
        return True, f"max_failed_attempts_reached={failed_attempts}"

    attempted_at = _parse_iso_timestamp(str(existing_status.get("attempted_at") or ""))
    if attempted_at is None:
        attempted_at = _parse_iso_timestamp(str(existing_status.get("sent_at") or ""))
    if attempted_at is None:
        return False, ""

    retry_after = attempted_at + timedelta(hours=FAILED_RETRY_COOLDOWN_HOURS)
    now = datetime.now(tz=attempted_at.tzinfo) if attempted_at.tzinfo else datetime.utcnow()
    if now < retry_after:
        return True, f"retry_cooldown_until={retry_after.isoformat()}"
    return False, ""


def load_leads(city_slug: str) -> list[dict[str, Any]]:
    slug_variants = [city_slug]
    if "-" in city_slug:
        slug_variants.append(city_slug.replace("-", "_"))
    if "_" in city_slug:
        slug_variants.append(city_slug.replace("_", "-"))

    candidate_paths: list[Path] = []
    seen: set[Path] = set()

    def _add(path: Path) -> None:
        resolved = path.resolve()
        if resolved in seen:
            return
        seen.add(resolved)
        candidate_paths.append(path)

    for data_root in _candidate_public_data_roots():
        for slug in slug_variants:
            _add(data_root / slug / f"{slug}_leads.json")
            for path in sorted(data_root.glob(f"*/{slug}/{slug}_leads.json")):
                _add(path)
            # Backward compatibility with legacy export location.
            _add(data_root / f"{slug}_leads.json")

    leads_path = next((path for path in candidate_paths if path.exists()), None)
    if leads_path is None:
        searched = ", ".join(str(path) for path in candidate_paths)
        raise FileNotFoundError(f"Leads file not found. Searched: {searched}")

    payload = json.loads(leads_path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError("Leads file must contain a JSON array.")
    return [item for item in payload if isinstance(item, dict)]


def load_leads_auto(
    city_slug: str | None = None,
    *,
    allow_json_fallback_when_excel_empty: bool = False,
) -> list[dict[str, Any]]:
    """
    Load leads from output Excel files (primary) or JSON (fallback).
    If city_slug is provided and Excel files exist, filter to that city.
    """
    try:
        from .excel_leads import find_output_excel_files, load_all_excel_leads
    except ImportError:
        from excel_leads import find_output_excel_files, load_all_excel_leads

    excel_files = find_output_excel_files(OUTPUT_DIR)
    if excel_files:
        logger.info("Loading leads from %d Excel file(s) in output/", len(excel_files))
        leads = load_all_excel_leads(OUTPUT_DIR)
        if city_slug:
            leads = [
                l for l in leads
                if l.get("city", "").strip().lower().replace(" ", "-") == city_slug
                or l.get("location", "").strip().lower().replace(" ", "-") == city_slug
            ]
        if leads:
            return leads
        if not allow_json_fallback_when_excel_empty:
            logger.info(
                "Excel leads exist but no eligible approved/deployed rows were found for city=%s.",
                city_slug or "",
            )
            return []
        logger.info("Excel leads found no eligible rows. Falling back to JSON leads.")

    if not excel_files:
        logger.info("No output Excel files found. Falling back to JSON leads.")
    if city_slug is None:
        raise SystemExit(
            "No output Excel files found and no city_slug provided "
            "for JSON fallback. Run website-builder first."
        )
    return load_leads(city_slug)


def _slugify_optional(value: str) -> str:
    text = re.sub(r"[^a-z0-9]+", "-", str(value or "").strip().lower()).strip("-")
    return text


def _excel_header_map(ws: Any) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col).value or "").strip().lower()
        if header:
            result[header] = col
    return result


def _excel_cell(ws: Any, row_idx: int, header_map: dict[str, int], key: str, default: str = "") -> str:
    col = header_map.get(key.lower())
    if col is None:
        return default
    return str(ws.cell(row_idx, col).value or "").strip()


def _workbook_city_slug(file_path: Path) -> str:
    try:
        rel = file_path.resolve().relative_to(OUTPUT_DIR.resolve())
    except ValueError:
        return ""
    if len(rel.parts) >= 3:
        return _slugify_optional(rel.parts[1])
    return ""


def _stable_lead_id(file_path: Path, row_idx: int, seed: str) -> str:
    digest = hashlib.sha1(f"{file_path}::{row_idx}::{seed}".encode("utf-8")).hexdigest()[:10].upper()
    return f"LEAD_{digest}"


def load_output_excel_leads(
    output_root: Path,
    city_slug: str | None = None,
) -> tuple[list[dict[str, Any]], dict[Path, dict[str, Any]]]:
    excel_files = sorted(output_root.rglob("leads.xlsx"))
    contexts: dict[Path, dict[str, Any]] = {}
    leads: list[dict[str, Any]] = []

    for file_path in excel_files:
        workbook_city = _workbook_city_slug(file_path)
        if city_slug and workbook_city and workbook_city != city_slug:
            continue

        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        header_map = _excel_header_map(worksheet)
        contexts[file_path.resolve()] = {
            "workbook": workbook,
            "worksheet": worksheet,
            "header_map": header_map,
            "dirty": False,
        }

        for row_idx in range(2, worksheet.max_row + 1):
            row_city = _excel_cell(worksheet, row_idx, header_map, "city")
            row_city_slug = _slugify_optional(row_city)
            if city_slug and row_city_slug and row_city_slug != city_slug:
                continue

            email = _excel_cell(worksheet, row_idx, header_map, "email")
            shop_id = _excel_cell(worksheet, row_idx, header_map, "shop_id")
            source_website = _excel_cell(worksheet, row_idx, header_map, "source_website")
            generated_website = _excel_cell(worksheet, row_idx, header_map, "generated_website")
            website_url = _excel_cell(worksheet, row_idx, header_map, "website_url")
            effective_website_url = website_url or generated_website or source_website
            lead_id_seed = shop_id or email or _excel_cell(worksheet, row_idx, header_map, "name")
            lead_id = shop_id or _stable_lead_id(file_path, row_idx, lead_id_seed)

            lead: dict[str, Any] = {
                "lead_id": lead_id,
                "shop_name": _excel_cell(worksheet, row_idx, header_map, "name"),
                "primary_email": email,
                "emails": [email] if email else [],
                "phone": _excel_cell(worksheet, row_idx, header_map, "phone"),
                "category": _excel_cell(worksheet, row_idx, header_map, "category"),
                "location": row_city,
                "city": row_city,
                "country": _excel_cell(worksheet, row_idx, header_map, "country"),
                "website_url": effective_website_url,
                "generated_website": generated_website,
                "source_website": source_website,
                "whatsapp": _excel_cell(worksheet, row_idx, header_map, "whatsapp"),
                "review_status": _excel_cell(worksheet, row_idx, header_map, "review_status"),
                "email_status": _excel_cell(worksheet, row_idx, header_map, "email_status"),
                "rating": _excel_cell(worksheet, row_idx, header_map, "rating"),
                "review_count": _excel_cell(worksheet, row_idx, header_map, "review_count"),
                "_source_file": str(file_path.resolve()),
                "_source_row": row_idx,
            }
            leads.append(lead)

    return leads, contexts


def _ensure_excel_column(context: dict[str, Any], header: str) -> int:
    header_map: dict[str, int] = context["header_map"]
    worksheet = context["worksheet"]
    existing = header_map.get(header.lower())
    if existing is not None:
        return existing
    col_idx = worksheet.max_column + 1
    worksheet.cell(1, col_idx).value = header
    header_map[header.lower()] = col_idx
    context["dirty"] = True
    return col_idx


def _format_email_status(status: str, reason: str = "") -> str:
    normalized_status = status.strip().lower()
    normalized_reason = reason.strip()
    if normalized_reason:
        return f"{normalized_status}: {normalized_reason}"
    return normalized_status


def update_excel_email_status(
    contexts: dict[Path, dict[str, Any]],
    lead: dict[str, Any],
    status: str,
    reason: str = "",
) -> None:
    source_file_raw = str(lead.get("_source_file") or "").strip()
    source_row = int(lead.get("_source_row") or 0)
    if not source_file_raw or source_row <= 1:
        return

    source_path = Path(source_file_raw).resolve()
    context = contexts.get(source_path)
    if context is None:
        return

    worksheet = context["worksheet"]
    email_status_col = _ensure_excel_column(context, "email_status")
    worksheet.cell(source_row, email_status_col).value = _format_email_status(status, reason)
    context["dirty"] = True


def save_excel_contexts(contexts: dict[Path, dict[str, Any]], *, dry_run: bool) -> None:
    for path, context in contexts.items():
        workbook = context["workbook"]
        try:
            if context.get("dirty") and not dry_run:
                workbook.save(path)
        finally:
            workbook.close()


def load_status(city_slug: str) -> dict[str, dict[str, Any]]:
    status_path = _email_status_root() / f"{city_slug}.json"
    if not status_path.exists():
        return {}
    try:
        payload = json.loads(status_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    if not isinstance(payload, dict):
        return {}
    return {k: v for k, v in payload.items() if isinstance(k, str) and isinstance(v, dict)}


def save_status(city_slug: str, status_data: dict[str, dict[str, Any]]) -> None:
    status_path = _email_status_root() / f"{city_slug}.json"
    status_path.parent.mkdir(parents=True, exist_ok=True)
    status_path.write_text(json.dumps(status_data, indent=2, ensure_ascii=False), encoding="utf-8")


def load_dedup_store(city_slug: str) -> dict[str, dict[str, Any]]:
    dedup_path = _email_status_root() / f"{city_slug}_dedup.json"
    if not dedup_path.exists():
        return {}
    try:
        payload = json.loads(dedup_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    if not isinstance(payload, dict):
        return {}
    return {k: v for k, v in payload.items() if isinstance(k, str) and isinstance(v, dict)}


def save_dedup_store(city_slug: str, dedup: dict[str, dict[str, Any]]) -> None:
    dedup_path = _email_status_root() / f"{city_slug}_dedup.json"
    dedup_path.parent.mkdir(parents=True, exist_ok=True)
    dedup_path.write_text(json.dumps(dedup, indent=2, ensure_ascii=False), encoding="utf-8")


def update_daily_send_summary(summary_path: Path, timestamp_iso: str) -> None:
    date_key = timestamp_iso[:10]
    payload: dict[str, Any] = {}
    if summary_path.exists():
        try:
            loaded = json.loads(summary_path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                payload = loaded
        except Exception:
            payload = {}
    current = payload.get(date_key, 0)
    if not isinstance(current, int):
        current = 0
    payload[date_key] = current + 1
    summary_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def normalize_email(email: str) -> str:
    return email.strip().lower()


def resolve_lead_email(lead: dict[str, Any]) -> str:
    primary_email = lead.get("primary_email")
    if isinstance(primary_email, str) and primary_email.strip():
        return primary_email.strip()

    emails = lead.get("emails")
    if isinstance(emails, list):
        for value in emails:
            if isinstance(value, str) and value.strip():
                return value.strip()
    return ""


def _normalize_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def _extract_subject_and_body(content: str) -> tuple[str, str]:
    subject_match = re.search(r"(?im)^\s*subject\s*:\s*(.+)$", content)
    body_match = re.search(r"(?ims)^\s*body\s*:\s*(.+)$", content)

    subject = subject_match.group(1).strip() if subject_match else ""
    body = body_match.group(1).strip() if body_match else ""

    if not body:
        lines = [line.rstrip() for line in content.splitlines() if line.strip()]
        if lines:
            if lines[0].lower().startswith("subject:"):
                subject = subject or lines[0].split(":", 1)[1].strip()
                body = "\n".join(lines[1:]).strip()
            else:
                body = "\n".join(lines).strip()

    if not subject:
        subject = "Quick idea for your business website"
    if not body:
        body = "I have a quick idea to help your business attract more local customers online."
    return subject, body


def _extract_json_object(raw: str) -> dict[str, Any]:
    text = str(raw or "").strip()
    if not text:
        raise ValueError("Empty JSON response")
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError("No JSON object found")
    parsed = json.loads(text[start : end + 1])
    if not isinstance(parsed, dict):
        raise ValueError("JSON response is not an object")
    return parsed


def _website_scenario(lead: dict[str, Any]) -> str:
    source_website = str(lead.get("source_website") or lead.get("website_url") or "").strip()
    generated_website = str(lead.get("generated_website") or "").strip()
    weakness_summary = str(lead.get("weakness_summary") or "").strip()

    if not source_website:
        return "No website present"
    if weakness_summary:
        return weakness_summary
    if generated_website:
        return f"Existing site present; generated replacement preview available at {generated_website}"
    return f"Existing website present at {source_website}"


def generate_email_nvidia(lead: dict, template: str) -> dict:
    if not NVIDIA_EMAIL_API_KEY:
        raise RuntimeError("Missing NVIDIA_EMAIL_API_KEY / NVIDIA_API_KEY environment variable.")

    try:
        from openai import OpenAI
    except ImportError as exc:
        raise RuntimeError("openai package is required for NVIDIA generation.") from exc

    client = OpenAI(base_url=NVIDIA_API_BASE, api_key=NVIDIA_EMAIL_API_KEY)
    system_prompt = (
        "You are an expert cold email copywriter for a web development agency called "
        "NexviaTech targeting small and medium businesses in India. Write concise, "
        "personalized, conversion-focused cold emails. Never sound like a template. "
        "Always reference something specific about the business."
    )
    user_prompt = (
        "Write a cold outreach email for this lead. Return JSON with fields: subject, "
        "body (plain text, max 180 words), ps_line (one punchy P.S. line).\n\n"
        f"Business: {str(lead.get('shop_name') or lead.get('name') or '').strip()}\n"
        f"Category: {str(lead.get('category') or '').strip()}\n"
        f"City: {str(lead.get('city') or lead.get('location') or '').strip()}\n"
        f"Website situation: {_website_scenario(lead)}\n"
        f"AI weakness summary: {str(lead.get('weakness_summary') or '').strip()}\n"
        f"Outreach angle: {str(lead.get('outreach_angle') or '').strip()}\n"
        f"Base template hints: {template}"
    )

    response = client.chat.completions.create(
        model=MODEL_CONFIG["email"],
        temperature=0.4,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    )
    content = response.choices[0].message.content if response.choices else ""
    payload = _extract_json_object(content or "")

    subject = str(payload["subject"]).strip()[:120]
    body = _word_trim(str(payload["body"]).strip(), 180)
    ps_line = str(payload["ps_line"]).strip()
    if not subject or not body or not ps_line:
        raise ValueError("Missing required email fields from NVIDIA response")
    return {
        "subject": subject,
        "body": body,
        "ps_line": ps_line,
    }


def _word_trim(text: str, max_words: int) -> str:
    words = text.split()
    if len(words) <= max_words:
        return text.strip()
    return " ".join(words[:max_words]).strip()


def _to_float_or_none(value: Any) -> float | None:
    try:
        if value is None:
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_int_or_none(value: Any) -> int | None:
    try:
        if value is None:
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def resolve_bucket_and_scenario(
    category: str,
    category_bucket_doc: dict[str, Any],
) -> tuple[str, str, str]:
    normalized_category = _normalize_text(category)

    bucket_key = "default"
    categories = category_bucket_doc.get("categories", {})
    if isinstance(categories, dict):
        for key, value in categories.items():
            if isinstance(value, list):
                if any(_normalize_text(str(item)) == normalized_category for item in value):
                    bucket_key = str(key)
                    break
            elif isinstance(value, str):
                if _normalize_text(value) == normalized_category:
                    bucket_key = str(key)
                    break

    scenarios = category_bucket_doc.get("scenarios", {})
    scenario = "default"
    if isinstance(scenarios, dict):
        candidate = scenarios.get(bucket_key, "default")
        if isinstance(candidate, str):
            scenario = candidate

    bucket_no_map = category_bucket_doc.get("bucket_no", {})
    bucket_no = "0"
    if isinstance(bucket_no_map, dict):
        raw_no = bucket_no_map.get(bucket_key, 0)
        bucket_no = str(raw_no)

    return bucket_key, scenario, bucket_no


def render_template(template: str, context: dict[str, str]) -> str:
    rendered = template
    for key, value in context.items():
        rendered = rendered.replace(f"{{{{{key}}}}}", value)
    return rendered


def _template_entry_to_text(value: Any) -> str:
    if isinstance(value, str):
        return value
    if not isinstance(value, dict):
        return ""

    subject = str(value.get("subject") or "").strip()
    body = str(value.get("body") or "").strip()
    parts: list[str] = []
    if subject:
        parts.append(f"Subject guidance: {subject}")
    if body:
        parts.append(f"Body guidance:\n{body}")
    return "\n\n".join(parts).strip()


def select_template(
    bucket_key: str,
    scenario: str,
    bucket_template_doc: dict[str, Any],
) -> str:
    templates = bucket_template_doc.get("templates", {})
    if not isinstance(templates, dict):
        return ""
    bucket_templates = templates.get(bucket_key)
    direct = _template_entry_to_text(bucket_templates)
    if direct:
        return direct
    if isinstance(bucket_templates, dict):
        scenario_template = _template_entry_to_text(bucket_templates.get(scenario))
        if scenario_template:
            return scenario_template
        default_template = _template_entry_to_text(bucket_templates.get("default"))
        if default_template:
            return default_template
    return ""


def detect_pii(text: str) -> None:
    if not NVIDIA_GUARD_API_KEY:
        return
    payload = {
        "model": MODEL_CONFIG["pii"],
        "messages": [{"role": "user", "content": f"Identify PII in: {text}"}],
        "temperature": 0.1,
        "max_tokens": 16,
    }
    headers = {
        "Authorization": f"Bearer {NVIDIA_GUARD_API_KEY}",
        "Content-Type": "application/json",
    }
    try:
        resp = requests.post(
            f"{NVIDIA_API_BASE}/chat/completions",
            headers=headers,
            json=payload,
            timeout=(10, 45),
        )
        resp.raise_for_status()
    except Exception as exc:
        logger.warning("PII detection failed, continuing")


def is_email_safe(subject: str, body: str) -> bool:
    if not NVIDIA_GUARD_API_KEY:
        return True
    prompt = (
        "Review this cold sales email for spam signals, aggressive language, "
        "misleading claims, or content that could get the sender domain blacklisted. "
        "Reply with one word only: SAFE or UNSAFE.\n\n"
        f"Subject: {subject}\n"
        f"Body: {body}"
    )
    payload = {
        "model": MODEL_CONFIG["safety"],
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.3,
        "max_tokens": 16,
    }
    headers = {
        "Authorization": f"Bearer {NVIDIA_GUARD_API_KEY}",
        "Content-Type": "application/json",
    }
    try:
        resp = requests.post(
            f"{NVIDIA_API_BASE}/chat/completions",
            headers=headers,
            json=payload,
            timeout=(10, 45),
        )
        if resp.status_code != 200:
            logger.warning("NVIDIA Guardrail returned HTTP %d, blocking send.", resp.status_code)
            return False
        content = resp.json()["choices"][0]["message"]["content"].strip().upper()
        return content == "SAFE"
    except Exception as exc:
        logger.warning("NVIDIA Guardrail check failed (blocking send): %s", exc)
        return False


def send_email_via_zoho(to_email: str, subject: str, body: str) -> tuple[str, str]:
    password = SMTP_PASSWORD
    if not password:
        raise RuntimeError("Missing EMAIL_HOST_PASSWORD environment variable.")

    msg_id = make_msgid()
    message = EmailMessage()
    message["From"] = SMTP_USERNAME
    message["To"] = to_email
    message["Subject"] = subject
    message["Message-ID"] = msg_id
    message.set_content(body)

    smtp_response = "accepted"

    def _send() -> str:
        nonlocal smtp_response
        security_mode = SMTP_SECURITY if SMTP_SECURITY in {"ssl", "starttls", "plain"} else "ssl"
        if security_mode == "ssl":
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=45) as smtp:
                smtp.login(SMTP_USERNAME, password)
                refused = smtp.send_message(message)
                smtp_response = "accepted" if not refused else json.dumps(refused)
                return smtp_response
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=45) as smtp:
            smtp.ehlo()
            if security_mode == "starttls":
                smtp.starttls()
                smtp.ehlo()
            smtp.login(SMTP_USERNAME, password)
            refused = smtp.send_message(message)
            smtp_response = "accepted" if not refused else json.dumps(refused)
            return smtp_response

    provider_response = retry_operation("smtp_send", _send, logger)
    return msg_id, provider_response


def update_status(
    status_data: dict[str, dict[str, Any]],
    lead_id: str,
    subject: str,
    status: str,
    timestamp_iso: str,
) -> None:
    current = status_data.get(lead_id, {})
    if not isinstance(current, dict):
        current = {}
    previous_failed = _to_int_or_none(current.get("failed_attempts")) or 0
    status_data[lead_id] = {
        "email_sent": status == "SENT",
        "attempted_at": timestamp_iso,
        "sent_at": timestamp_iso if status == "SENT" else current.get("sent_at"),
        "subject": subject,
        "status": status,
        "failed_attempts": 0 if status == "SENT" else previous_failed + 1,
    }


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument("city_name", nargs="?", default=None)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--require-approved-review",
        action="store_true",
        help='Abort if no row with review_status="approved" or "good" exists in any output leads.xlsx file.',
    )
    parser.add_argument(
        "--dry-run-no-gen",
        action="store_true",
        help="Only print eligible recipients; skip generation and SMTP send.",
    )
    return parser


def main() -> None:
    _configure_logging()
    args = _build_parser().parse_args()
    _ensure_template_dependencies_exist()
    if args.dry_run_no_gen and not args.dry_run:
        raise SystemExit("--dry-run-no-gen requires --dry-run.")

    if args.require_approved_review:
        approved_rows = _count_approved_review_rows(OUTPUT_DIR)
        if approved_rows == 0:
            raise SystemExit(
                "No approved/good review rows found in any leads.xlsx under output/. "
                'Run the website checker so at least one row has review_status="good" before outreach.'
            )

    if args.city_name:
        city_name = args.city_name.strip()
        if not city_name:
            raise SystemExit("City name cannot be empty.")
        try:
            city_slug = _sanitize_city_name(city_name)
        except ValueError as exc:
            raise SystemExit(str(exc)) from exc
    else:
        city_name = None
        city_slug = None

    effective_slug = city_slug or "output_excel"
    signature = _build_signature()
    needs_generation = not args.dry_run_no_gen
    needs_send = not args.dry_run
    _validate_required_config(
        signature=signature,
        needs_generation=needs_generation,
        needs_send=needs_send,
    )
    suppression_list = _load_suppression_list(SUPPRESSION_LIST_PATH)
    opt_out_footer = UNSUBSCRIBE_TEXT
    sent_count = 0
    skipped_count = 0
    leads, excel_contexts = load_output_excel_leads(OUTPUT_DIR, city_slug if city_slug else None)
    if not leads:
        save_excel_contexts(excel_contexts, dry_run=True)
        logger.info(
            "No leads.xlsx rows found for city=%s. Skipping email flow.",
            city_slug or "",
        )
        print("PIPELINE_STAT: emails_sent=0")
        print("PIPELINE_STAT: emails_skipped=0")
        return

    category_bucket_doc: dict[str, Any] = {}
    bucket_template_doc: dict[str, Any] = {}
    if needs_generation:
        try:
            category_bucket_doc, bucket_template_doc = validate_template_files(
                CATEGORY_BUCKET_PATH, BUCKET_TEMPLATE_PATH
            )
        except TemplateValidationError as exc:
            logger.error("Template validation failed: %s", exc)
            raise SystemExit(1) from exc

    status_data = load_status(effective_slug)
    dedup_store = load_dedup_store(effective_slug)

    email_status_dir = _email_status_root()
    smtp_log_path = email_status_dir / f"{effective_slug}_smtp_events.jsonl"
    audit_log_path = email_status_dir / f"{effective_slug}_audit_log.json"
    rate_state_path = email_status_dir / f"{effective_slug}_rate_state.json"
    limiter = RateLimiter(MAX_PER_HOUR, MAX_PER_DAY, rate_state_path)
    try:
        for lead in leads:
            lead_id = str(lead.get("lead_id") or "").strip()
            lead_label = _lead_event_label(lead)
            if not lead_id:
                skipped_count += 1
                update_excel_email_status(excel_contexts, lead, "skipped", "missing lead_id")
                _emit_pipeline_event("site", "unknown", "SKIPPED", "missing lead_id")
                continue

            review_status = str(lead.get("review_status") or "").strip().lower()
            if review_status not in {"approved", "good"}:
                skipped_count += 1
                update_excel_email_status(
                    excel_contexts,
                    lead,
                    "skipped",
                    f"review_status={review_status or 'blank'}",
                )
                _emit_pipeline_event("site", lead_label, "SKIPPED", f"review_status={review_status or 'blank'}")
                continue

            email = resolve_lead_email(lead)
            if not email:
                skipped_count += 1
                update_excel_email_status(excel_contexts, lead, "skipped", "missing email")
                _emit_pipeline_event("site", lead_label, "SKIPPED", "missing email")
                continue

            normalized_email = normalize_email(email)
            if not _is_valid_email(normalized_email):
                logger.warning("Skipping lead with invalid email format lead=%s email=%s", lead_id, email)
                skipped_count += 1
                update_excel_email_status(excel_contexts, lead, "skipped", "invalid email format")
                _emit_pipeline_event("site", lead_label, "SKIPPED", "invalid email format")
                continue

            suppressed, suppression_reason = _is_suppressed_recipient(normalized_email, suppression_list)
            if suppressed:
                logger.info("Skipping suppressed recipient lead=%s email=%s", lead_id, normalized_email)
                skipped_count += 1
                update_excel_email_status(excel_contexts, lead, "skipped", suppression_reason)
                _emit_pipeline_event("site", lead_label, "SKIPPED", suppression_reason)
                continue

            campaign_id = str(lead.get("campaign_id") or effective_slug).strip()
            dedup_key = f"{campaign_id}::{normalized_email}"

            existing_status = status_data.get(lead_id, {})
            if isinstance(existing_status, dict) and existing_status.get("email_sent") is True:
                update_excel_email_status(excel_contexts, lead, "sent")
                _emit_pipeline_event("site", lead_label, "SKIPPED", "already sent")
                continue

            should_skip, skip_reason = _should_skip_due_to_failures(existing_status)
            if should_skip:
                logger.info("Skipping lead due to failure policy lead=%s reason=%s", lead_id, skip_reason)
                skipped_count += 1
                update_excel_email_status(excel_contexts, lead, "skipped", skip_reason)
                _emit_pipeline_event("site", lead_label, "SKIPPED", skip_reason)
                continue

            if dedup_key in dedup_store:
                logger.info("Skipping duplicate recipient in same campaign: %s", dedup_key)
                skipped_count += 1
                update_excel_email_status(excel_contexts, lead, "skipped", "duplicate recipient in campaign")
                _emit_pipeline_event("site", lead_label, "SKIPPED", "duplicate recipient in campaign")
                continue

            if needs_send:
                allowed, reason = limiter.can_send()
                if not allowed:
                    logger.info("Stopping send flow: %s", reason)
                    skipped_count += 1
                    update_excel_email_status(excel_contexts, lead, "skipped", reason)
                    _emit_pipeline_event("site", lead_label, "SKIPPED", reason)
                    break

            bucket_key = "na"
            scenario = "na"
            bucket_no = "0"
            subject = ""
            body = ""
            if needs_generation:
                bucket_key, scenario, bucket_no = resolve_bucket_and_scenario(
                    str(lead.get("category") or ""),
                    category_bucket_doc,
                )
                template_text = select_template(bucket_key, scenario, bucket_template_doc)
                if not template_text:
                    logger.warning("No template found for bucket=%s scenario=%s", bucket_key, scenario)
                    skipped_count += 1
                    update_excel_email_status(excel_contexts, lead, "skipped", "no template available")
                    _emit_pipeline_event("site", lead_label, "SKIPPED", "no template available")
                    continue

                valid = False
                guardrail_reason = ""
                for _ in range(2):
                    try:
                        nvidia_email = generate_email_nvidia(lead, template_text)
                        subject = str(nvidia_email["subject"]).strip()
                        core_body = str(nvidia_email["body"]).strip()
                        ps_line = str(nvidia_email["ps_line"]).strip()
                        final_body = core_body
                        if ps_line:
                            final_body = f"{final_body}\n\n{ps_line}".strip()
                        final_body = f"{final_body}\n\n{signature}".strip()
                        if opt_out_footer:
                            final_body = f"{final_body}\n\n{opt_out_footer}".strip()
                        body = final_body
                    except Exception as exc:  # noqa: BLE001
                        logger.warning(
                            "NVIDIA generation failed for lead=%s error=%s. Falling back to Groq.",
                            lead_id,
                            exc,
                        )
                        guardrail_reason = f"generation failed: {exc}"
                        break
                    valid, guardrail_reason = validate_generated_email(subject, body, signature)
                    if valid:
                        break
                if not valid:
                    logger.warning(
                        "Skipping lead due to guardrail failure lead=%s reason=%s",
                        lead_id,
                        guardrail_reason or "generation_failed",
                    )
                    skipped_count += 1
                    update_excel_email_status(
                        excel_contexts,
                        lead,
                        "skipped",
                        guardrail_reason or "generation failed",
                    )
                    _emit_pipeline_event("site", lead_label, "SKIPPED", guardrail_reason or "generation failed")
                    continue

            if args.dry_run:
                if args.dry_run_no_gen:
                    logger.info("[DRY-RUN-NO-GEN] lead_id=%s email=%s", lead_id, normalized_email)
                else:
                    logger.info(
                        "[DRY-RUN] lead_id=%s email=%s bucket=%s scenario=%s subject=%s",
                        lead_id,
                        normalized_email,
                        bucket_no,
                        scenario,
                        subject,
                    )
                skipped_count += 1
                update_excel_email_status(excel_contexts, lead, "skipped", "dry run")
                _emit_pipeline_event("site", lead_label, "SKIPPED", "dry run")
                continue

            detect_pii(f"{subject}\n{body}")

            if not is_email_safe(subject, body):
                skipped_count += 1
                update_excel_email_status(excel_contexts, lead, "skipped", "safety_check_failed")
                _emit_pipeline_event("site", lead_label, "SKIPPED", "safety_check_failed")
                append_audit_log(
                    audit_log_path,
                    {"lead_id": lead_id, "skipped_reason": "safety_check_failed"},
                )
                continue

            try:
                msg_id, provider_response = send_email_via_zoho(normalized_email, subject, body)
                status_value = "SENT"
                limiter.record_send()
                sent_count += 1
                update_excel_email_status(excel_contexts, lead, "sent")
                _emit_pipeline_event("site", lead_label, "SENT")
            except Exception as exc:  # noqa: BLE001
                msg_id = ""
                provider_response = str(exc)
                status_value = "FAILED"
                skipped_count += 1
                update_excel_email_status(excel_contexts, lead, "failed", provider_response)
                _emit_pipeline_event("site", lead_label, "FAILED", provider_response)

            timestamp = utc_now_iso()
            smtp_row = {
                "message_id": msg_id,
                "provider_response": provider_response,
                "timestamp": timestamp,
                "recipient": normalized_email,
                "subject": subject,
            }
            append_jsonl(smtp_log_path, smtp_row)

            audit_row = {
                "lead_id": lead_id,
                "normalized_email": normalized_email,
                "bucket_no": bucket_no,
                "scenario": scenario,
                "subject": subject,
                "message_id": msg_id,
                "provider_response": provider_response,
                "status": status_value,
                "timestamp": timestamp,
                "whatsapp": str(lead.get("whatsapp") or "").strip(),
                "website_url": str(lead.get("website_url") or "").strip(),
            }
            append_audit_log(audit_log_path, audit_row)

            update_status(status_data, lead_id, subject, status_value, timestamp)
            save_status(effective_slug, status_data)

            if status_value == "SENT":
                dedup_store[dedup_key] = {
                    "lead_id": lead_id,
                    "email": normalized_email,
                    "campaign_id": campaign_id,
                    "timestamp": timestamp,
                }
                save_dedup_store(effective_slug, dedup_store)
                update_daily_send_summary(DAILY_SUMMARY_PATH, timestamp)
                RateLimiter.apply_random_delay()
    finally:
        save_excel_contexts(excel_contexts, dry_run=args.dry_run)

    print(f"PIPELINE_STAT: emails_sent={sent_count}")
    print(f"PIPELINE_STAT: emails_skipped={skipped_count}")


if __name__ == "__main__":
    main()
