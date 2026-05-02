from __future__ import annotations

import argparse
import json
import logging
import os
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


REPO_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(dotenv_path=REPO_ROOT / ".env")

OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", str(REPO_ROOT / "output"))).resolve()
PIPELINE_LOG_STRUCTURED = os.getenv("PIPELINE_LOG_FORMAT", "").strip().lower() == "structured"


def _emit_pipeline_event(entity: str, label: str, status: str, detail: str = "") -> bool:
    if not PIPELINE_LOG_STRUCTURED:
        return False
    payload = {
        "stage": "screenshot",
        "entity": entity,
        "label": label,
        "status": status,
    }
    if detail:
        payload["detail"] = detail
    print(f"PIPELINE_EVENT: {json.dumps(payload, ensure_ascii=False)}", flush=True)
    return True


def _configure_logging() -> None:
    if not logging.getLogger().handlers:
        logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Capture website preview screenshots and write screenshot_path into leads.xlsx.")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=OUTPUT_DIR,
        help=f"Root output directory to scan (default: {OUTPUT_DIR})",
    )
    parser.add_argument("--dry-run", action="store_true", help="Log planned screenshot work without writing files.")
    return parser


def _find_excel_files(output_dir: Path) -> list[Path]:
    return sorted(output_dir.rglob("leads.xlsx"))


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _ensure_column(ws: openpyxl.worksheet.worksheet.Worksheet, header_map: dict[str, int], header: str) -> int:
    existing = header_map.get(header.lower())
    if existing is not None:
        return existing
    next_col = ws.max_column + 1
    ws.cell(1, next_col).value = header
    header_map[header.lower()] = next_col
    return next_col


def _take_screenshot(url: str, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1400, "height": 900})
        try:
            page.goto(url, wait_until="networkidle", timeout=45000)
        except PlaywrightTimeoutError:
            page.goto(url, wait_until="load", timeout=45000)
            page.wait_for_timeout(2000)
        page.screenshot(path=str(output_path), full_page=False)
        browser.close()


def _process_workbook(file_path: Path, dry_run: bool) -> tuple[int, int]:
    logger = logging.getLogger("screenshot_taker")
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    header_map = {
        _normalize_header(worksheet.cell(1, col_idx).value): col_idx
        for col_idx in range(1, worksheet.max_column + 1)
        if _normalize_header(worksheet.cell(1, col_idx).value)
    }

    name_col = _ensure_column(worksheet, header_map, "name")
    shop_id_col = _ensure_column(worksheet, header_map, "shop_id")
    review_status_col = _ensure_column(worksheet, header_map, "review_status")
    website_url_col = _ensure_column(worksheet, header_map, "website_url")
    screenshot_col = _ensure_column(worksheet, header_map, "screenshot_path")

    captured = 0
    skipped = 0
    screenshot_dir = file_path.parent / "screenshots"

    for row_idx in range(2, worksheet.max_row + 1):
        review_status = str(worksheet.cell(row_idx, review_status_col).value or "").strip().lower()
        if review_status not in {"good", "approved"}:
            skipped += 1
            continue

        website_url = str(worksheet.cell(row_idx, website_url_col).value or "").strip()
        if not website_url.startswith(("http://", "https://")):
            skipped += 1
            continue

        shop_id = str(worksheet.cell(row_idx, shop_id_col).value or "").strip()
        name = str(worksheet.cell(row_idx, name_col).value or "").strip()
        screenshot_name = shop_id or name.lower().replace(" ", "-") or f"row-{row_idx}"
        screenshot_path = screenshot_dir / f"{screenshot_name}.png"

        if not dry_run and (not screenshot_path.exists() or not str(worksheet.cell(row_idx, screenshot_col).value or "").strip()):
            try:
                _take_screenshot(website_url, screenshot_path)
                captured += 1
            except Exception as exc:  # noqa: BLE001
                logger.warning("Failed screenshot for %s (%s): %s", name or shop_id, website_url, exc)
                skipped += 1
                continue
        elif dry_run:
            captured += 1

        worksheet.cell(row_idx, screenshot_col).value = str(screenshot_path.relative_to(REPO_ROOT)).replace("\\", "/")
        _emit_pipeline_event("site", shop_id or name or f"row-{row_idx}", "SUCCESS", str(screenshot_path))

    if not dry_run:
        workbook.save(file_path)
    workbook.close()
    return captured, skipped


def main() -> None:
    _configure_logging()
    args = _build_parser().parse_args()
    logger = logging.getLogger("screenshot_taker")

    files = _find_excel_files(args.output_dir.resolve())
    if not files:
        logger.info("No leads.xlsx files found under %s", args.output_dir)
        print("PIPELINE_STAT: screenshots_captured=0")
        return

    total_captured = 0
    total_skipped = 0
    for file_path in files:
        captured, skipped = _process_workbook(file_path, args.dry_run)
        total_captured += captured
        total_skipped += skipped
        logger.info("Processed %s captured=%d skipped=%d", file_path, captured, skipped)

    print(f"PIPELINE_STAT: screenshots_captured={total_captured}")
    print(f"PIPELINE_STAT: screenshots_skipped={total_skipped}")


if __name__ == "__main__":
    main()
