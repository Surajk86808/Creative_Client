from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from .checker import check_with_delay

logger = logging.getLogger("whatsappcheck")
WHATSAPP_COL_HEADER = "whatsapp"
PHONE_COL_HEADER = "phone"


def _find_col(headers: list[str], name: str) -> int | None:
    """Return 0-based column index or None."""
    normalized = [str(h or "").strip().lower() for h in headers]
    try:
        return normalized.index(name.lower())
    except ValueError:
        return None


def process_excel_file(file_path: Path, dry_run: bool = False) -> dict[str, int]:
    """
    Process a single leads.xlsx file.
    Returns stats: {checked, skipped, yes, no, invalid, error}
    """
    stats: dict[str, int] = {
        "checked": 0,
        "skipped": 0,
        "yes": 0,
        "no": 0,
        "invalid": 0,
        "error": 0,
    }

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Read header row (row 1)
    headers = [str(ws.cell(1, col).value or "") for col in range(1, ws.max_column + 1)]

    phone_idx = _find_col(headers, PHONE_COL_HEADER)
    if phone_idx is None:
        logger.warning("No 'phone' column in %s — skipping", file_path)
        return stats

    # Find or create whatsapp column
    wa_idx = _find_col(headers, WHATSAPP_COL_HEADER)
    if wa_idx is None:
        wa_idx = len(headers)
        ws.cell(1, wa_idx + 1).value = WHATSAPP_COL_HEADER
        headers.append(WHATSAPP_COL_HEADER)

    phone_col = phone_idx + 1  # openpyxl is 1-based
    wa_col = wa_idx + 1

    for row_idx in range(2, ws.max_row + 1):
        phone_raw = str(ws.cell(row_idx, phone_col).value or "").strip()
        existing_wa = str(ws.cell(row_idx, wa_col).value or "").strip()

        # Skip empty phones
        if not phone_raw:
            stats["skipped"] += 1
            continue

        # Idempotent: skip already-checked rows
        if existing_wa in ("YES", "NO", "INVALID", "ERROR"):
            stats["skipped"] += 1
            continue

        logger.info("Checking row %d phone=%s", row_idx, phone_raw)
        stats["checked"] += 1

        if not dry_run:
            _, result = check_with_delay(phone_raw)
            ws.cell(row_idx, wa_col).value = result
            stats[result.lower()] = stats.get(result.lower(), 0) + 1
        else:
            logger.info("[DRY-RUN] would check %s", phone_raw)

    if not dry_run:
        wb.save(file_path)
        logger.info("Saved %s", file_path)

    return stats


def find_all_excel_files(output_root: Path) -> list[Path]:
    """Recursively find all leads.xlsx files under output_root."""
    return sorted(output_root.rglob("leads.xlsx"))

