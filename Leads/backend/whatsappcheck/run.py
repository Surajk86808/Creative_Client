from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import time
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv


REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

load_dotenv(dotenv_path=REPO_ROOT / ".env")

OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", str(REPO_ROOT / "output"))).resolve()
DEFAULT_COUNTRY_CODE = re.sub(r"\D", "", os.getenv("WHATSAPP_DEFAULT_COUNTRY_CODE", "91")) or "91"
REQUEST_DELAY_SECONDS = float(os.getenv("WHATSAPP_REQUEST_DELAY", "2.0"))
WHATSAPP_COL_HEADER = "whatsapp"
PHONE_COL_HEADER = "phone"
REVIEW_STATUS_COL_HEADER = "review_status"
VALID_RESULTS = {"YES", "NO", "INVALID", "ERROR"}
PIPELINE_LOG_STRUCTURED = os.getenv("PIPELINE_LOG_FORMAT", "").strip().lower() == "structured"


def _emit_pipeline_event(entity: str, label: str, status: str, detail: str = "") -> bool:
    if not PIPELINE_LOG_STRUCTURED:
        return False
    payload = {
        "stage": "whatsapp",
        "entity": entity,
        "label": label,
        "status": status,
    }
    if detail:
        payload["detail"] = detail
    print(f"PIPELINE_EVENT: {json.dumps(payload, ensure_ascii=False)}", flush=True)
    return True


def _configure_logging() -> None:
    if not logging.getLogger().handlers:
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s %(levelname)s %(message)s",
        )


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Check WhatsApp availability for reviewed-good leads in output Excel files."
    )
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Path to a specific leads.xlsx file. If not set, all leads.xlsx files under OUTPUT_DIR are processed.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be checked without making requests or writing files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=OUTPUT_DIR,
        help=f"Root output directory to scan (default: {OUTPUT_DIR})",
    )
    return parser


def _find_all_excel_files(output_root: Path) -> list[Path]:
    return sorted(output_root.rglob("leads.xlsx"))


def _find_col(headers: list[str], name: str) -> int | None:
    normalized = [str(header or "").strip().lower() for header in headers]
    try:
        return normalized.index(name.lower())
    except ValueError:
        return None


def _normalize_phone_e164(raw: object, country_code: str = DEFAULT_COUNTRY_CODE) -> str | None:
    text = str(raw or "").strip()
    if not text:
        return None

    digits = re.sub(r"\D", "", text)
    if not digits:
        return None

    if text.startswith("+"):
        normalized_digits = digits
    elif digits.startswith("00"):
        normalized_digits = digits[2:]
    elif digits.startswith("0") and len(digits) > 10:
        normalized_digits = country_code + digits[1:]
    elif len(digits) == 10:
        normalized_digits = country_code + digits
    elif country_code and len(digits) < 11:
        normalized_digits = country_code + digits
    else:
        normalized_digits = digits

    if not normalized_digits.isdigit():
        return None
    if len(normalized_digits) < 8 or len(normalized_digits) > 15:
        return None

    return f"+{normalized_digits}"


def _check_whatsapp_head(phone_e164: str, timeout: int = 10) -> str:
    url = f"https://wa.me/{phone_e164.lstrip('+')}"
    try:
        response = requests.head(
            url,
            timeout=timeout,
            allow_redirects=False,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36"
                )
            },
        )
    except requests.RequestException:
        return "ERROR"

    status_code = int(response.status_code)
    if status_code in {301, 302, 303, 307, 308}:
        location = str(response.headers.get("Location") or "").lower()
        if "whatsapp" in location:
            return "YES"
        return "NO"
    if status_code in {200, 204}:
        return "YES"
    if status_code in {400, 404, 410}:
        return "NO"
    return "ERROR"


def _process_excel_file(file_path: Path, dry_run: bool = False) -> dict[str, int]:
    logger = logging.getLogger("whatsappcheck")
    stats: dict[str, int] = {
        "checked": 0,
        "skipped": 0,
        "yes": 0,
        "no": 0,
        "invalid": 0,
        "error": 0,
    }

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    headers = [str(worksheet.cell(1, col).value or "") for col in range(1, worksheet.max_column + 1)]

    phone_idx = _find_col(headers, PHONE_COL_HEADER)
    if phone_idx is None:
        logger.warning("No '%s' column in %s - skipping", PHONE_COL_HEADER, file_path)
        return stats

    review_status_idx = _find_col(headers, REVIEW_STATUS_COL_HEADER)
    if review_status_idx is None:
        logger.info("No '%s' column in %s - no eligible rows", REVIEW_STATUS_COL_HEADER, file_path)
        return stats

    whatsapp_idx = _find_col(headers, WHATSAPP_COL_HEADER)
    if whatsapp_idx is None:
        whatsapp_idx = len(headers)
        worksheet.cell(1, whatsapp_idx + 1).value = WHATSAPP_COL_HEADER
        headers.append(WHATSAPP_COL_HEADER)

    phone_col = phone_idx + 1
    review_status_col = review_status_idx + 1
    whatsapp_col = whatsapp_idx + 1
    changed = False

    for row_idx in range(2, worksheet.max_row + 1):
        review_status = str(worksheet.cell(row_idx, review_status_col).value or "").strip().lower()
        if review_status != "good":
            stats["skipped"] += 1
            continue

        existing_whatsapp = str(worksheet.cell(row_idx, whatsapp_col).value or "").strip().upper()
        if existing_whatsapp:
            stats["skipped"] += 1
            continue

        phone_raw = str(worksheet.cell(row_idx, phone_col).value or "").strip()
        if not phone_raw:
            stats["skipped"] += 1
            continue

        normalized_phone = _normalize_phone_e164(phone_raw)
        if normalized_phone is None:
            result = "INVALID"
        else:
            logger.info("Checking row %d phone=%s normalized=%s", row_idx, phone_raw, normalized_phone)
            if dry_run:
                logger.info("[DRY-RUN] would HEAD https://wa.me/%s", normalized_phone.lstrip("+"))
                stats["checked"] += 1
                continue
            time.sleep(REQUEST_DELAY_SECONDS)
            result = _check_whatsapp_head(normalized_phone)

        if result not in VALID_RESULTS:
            result = "ERROR"

        worksheet.cell(row_idx, whatsapp_col).value = result
        changed = True
        stats["checked"] += 1
        stats[result.lower()] = stats.get(result.lower(), 0) + 1

    if changed and not dry_run:
        workbook.save(file_path)
        logger.info("Saved %s", file_path)

    return stats


def main() -> None:
    _configure_logging()
    args = _build_parser().parse_args()
    logger = logging.getLogger("whatsappcheck")

    if args.file:
        files = [args.file.resolve()]
    else:
        files = _find_all_excel_files(args.output_dir)

    if not files:
        logger.info("No leads.xlsx files found under %s", args.output_dir)
        print("PIPELINE_STAT: whatsapp_checked=0")
        print("PIPELINE_STAT: whatsapp_yes=0")
        return

    total_stats: dict[str, int] = {
        "checked": 0,
        "skipped": 0,
        "yes": 0,
        "no": 0,
        "invalid": 0,
        "error": 0,
    }

    for file_path in files:
        _emit_pipeline_event("file", file_path.parent.name or file_path.name, "START")
        logger.info("Processing: %s", file_path)
        stats = _process_excel_file(file_path, dry_run=args.dry_run)
        for key, value in stats.items():
            total_stats[key] = total_stats.get(key, 0) + value
        _emit_pipeline_event(
            "file",
            file_path.parent.name or file_path.name,
            "SUCCESS",
            (
                f"checked={stats['checked']}, yes={stats['yes']}, no={stats['no']}, "
                f"invalid={stats['invalid']}, error={stats['error']}"
            ),
        )
        logger.info(
            "  -> checked=%d skipped=%d yes=%d no=%d invalid=%d error=%d",
            stats["checked"],
            stats["skipped"],
            stats["yes"],
            stats["no"],
            stats["invalid"],
            stats["error"],
        )

    logger.info(
        "TOTAL: checked=%d skipped=%d yes=%d no=%d invalid=%d error=%d",
        total_stats["checked"],
        total_stats["skipped"],
        total_stats["yes"],
        total_stats["no"],
        total_stats["invalid"],
        total_stats["error"],
    )
    print(f"PIPELINE_STAT: whatsapp_checked={total_stats['checked']}")
    print(f"PIPELINE_STAT: whatsapp_yes={total_stats['yes']}")


if __name__ == "__main__":
    main()
