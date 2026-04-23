from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
from dataclasses import dataclass
from datetime import UTC, datetime
from html import unescape
from html.parser import HTMLParser
from pathlib import Path

import openpyxl
from dotenv import load_dotenv


REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

load_dotenv(dotenv_path=REPO_ROOT / ".env")

OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", str(REPO_ROOT / "output"))).resolve()
SITE_SOURCE_SUFFIXES = {".html", ".tsx", ".jsx", ".ts", ".js", ".md", ".mdx"}
SKIP_DIR_NAMES = {"node_modules", "dist", ".git", "build", "coverage", "__pycache__"}
EMAIL_RE = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE)
PHONE_CANDIDATE_RE = re.compile(r"(?:\+?\d[\d\s().-]{7,}\d)")
WORD_RE = re.compile(r"[A-Za-z0-9']+")
STRING_LITERAL_RE = re.compile(
    r'"([^"\\]*(?:\\.[^"\\]*)*)"|\'([^\'\\]*(?:\\.[^\'\\]*)*)\'|`([^`$\\]*(?:\\.[^`\\]*)*)`',
    re.DOTALL,
)
JSX_TEXT_RE = re.compile(r">([^<>{}]+)<", re.DOTALL)
PLACEHOLDER_RULES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\blorem ipsum\b", re.IGNORECASE), "contains placeholder text (Lorem Ipsum)"),
    (re.compile(r"\bdolor sit amet\b", re.IGNORECASE), "contains placeholder text"),
    (re.compile(r"\[\[[A-Z0-9_]+\]\]"), "contains unreplaced template tokens"),
    (
        re.compile(r"\{\{\s*[A-Z_][A-Z0-9_ ]*\s*\}\}", re.IGNORECASE),
        "contains unreplaced template variables",
    ),
    (re.compile(r"\bexample@example\.com\b", re.IGNORECASE), "contains placeholder email"),
    (re.compile(r"\btest@example\.com\b", re.IGNORECASE), "contains placeholder email"),
    (
        re.compile(r"\b(?:123[-.\s]?456[-.\s]?7890|555[-.\s]?555[-.\s]?5555)\b"),
        "contains placeholder phone number",
    ),
]
PIPELINE_LOG_STRUCTURED = os.getenv("PIPELINE_LOG_FORMAT", "").strip().lower() == "structured"


def _emit_pipeline_event(entity: str, label: str, status: str, detail: str = "") -> bool:
    if not PIPELINE_LOG_STRUCTURED:
        return False
    payload = {
        "stage": "check",
        "entity": entity,
        "label": label,
        "status": status,
    }
    if detail:
        payload["detail"] = detail
    print(f"PIPELINE_EVENT: {json.dumps(payload, ensure_ascii=False)}", flush=True)
    return True


@dataclass
class SiteCheckResult:
    site_dir: Path
    index_path: Path
    meta_path: Path
    excel_path: Path
    shop_id: str
    meta: dict
    status: str
    reason: str
    local_file_url: str
    deployed_url: str = ""
    effective_url: str = ""
    title: str = ""
    visible_word_count: int = 0


class VisibleHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._skip_depth = 0
        self._in_title = False
        self.visible_chunks: list[str] = []
        self.title_chunks: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        lowered = tag.lower()
        if lowered in {"script", "style", "noscript", "template"}:
            self._skip_depth += 1
        if lowered == "title":
            self._in_title = True

    def handle_endtag(self, tag: str) -> None:
        lowered = tag.lower()
        if lowered in {"script", "style", "noscript", "template"} and self._skip_depth > 0:
            self._skip_depth -= 1
        if lowered == "title":
            self._in_title = False

    def handle_data(self, data: str) -> None:
        text = unescape(data or "").strip()
        if not text:
            return
        if self._in_title:
            self.title_chunks.append(text)
            return
        if self._skip_depth == 0:
            self.visible_chunks.append(text)


def _utc_now_iso() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _configure_logging() -> None:
    if not logging.getLogger().handlers:
        logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Check generated website folders and write review results to metadata and leads.xlsx."
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=OUTPUT_DIR,
        help=f"Root output directory to scan (default: {OUTPUT_DIR})",
    )
    parser.add_argument(
        "--site-dir",
        type=Path,
        default=None,
        help="Optional specific site directory under output/{country}/{city}/{category}/{place_id}.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would change without writing metadata or Excel files.",
    )
    return parser


def _iter_site_dirs(output_dir: Path) -> list[Path]:
    if not output_dir.exists():
        return []

    discovered: list[Path] = []
    seen: set[Path] = set()
    for meta_path in sorted(output_dir.rglob("_lead_meta.json")):
        site_dir = meta_path.parent
        if any(part in SKIP_DIR_NAMES for part in site_dir.parts):
            continue
        if not (site_dir / "index.html").exists():
            continue
        resolved = site_dir.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        discovered.append(site_dir)
    return discovered


def _looks_like_phone(text: str) -> bool:
    digits = re.sub(r"\D", "", text or "")
    return 10 <= len(digits) <= 15


def _has_phone_or_email(raw_text: str) -> bool:
    if EMAIL_RE.search(raw_text):
        return True
    return any(_looks_like_phone(match.group(0)) for match in PHONE_CANDIDATE_RE.finditer(raw_text))


def _iter_site_source_files(site_dir: Path) -> list[Path]:
    files: list[Path] = []
    index_path = site_dir / "index.html"
    if index_path.exists():
        files.append(index_path)

    for path in sorted(site_dir.rglob("*")):
        if not path.is_file():
            continue
        if any(part in SKIP_DIR_NAMES for part in path.parts):
            continue
        if path == index_path:
            continue
        if path.suffix.lower() in SITE_SOURCE_SUFFIXES:
            files.append(path)
    return files


def _is_visible_candidate(value: str) -> bool:
    text = unescape(value or "").strip()
    if not text:
        return False
    if text.startswith(("http://", "https://", "#", "/", "./", "../")):
        return False
    if text.endswith((".png", ".jpg", ".jpeg", ".svg", ".webp", ".tsx", ".ts", ".js", ".json")):
        return False
    if any(token in text for token in ("className", "bg-", "text-", "hover:", "transition-", "rgba(", "calc(")):
        return False
    if re.fullmatch(r"[A-Za-z0-9_./:#?&=%+-]+", text) and " " not in text:
        return False
    if re.fullmatch(r"[0-9 .()+-]+", text):
        return False
    return any(ch.isalpha() for ch in text)


def _extract_text_from_html(text: str) -> tuple[str, str]:
    parser = VisibleHTMLParser()
    parser.feed(text)
    title = " ".join(parser.title_chunks).strip()
    visible = " ".join(parser.visible_chunks).strip()
    return title, visible


def _extract_candidate_text_from_source(text: str) -> str:
    chunks: list[str] = []

    for match in JSX_TEXT_RE.finditer(text):
        candidate = unescape(match.group(1)).strip()
        if _is_visible_candidate(candidate):
            chunks.append(candidate)

    for match in STRING_LITERAL_RE.finditer(text):
        candidate = next((group for group in match.groups() if group is not None), "")
        candidate = bytes(candidate, "utf-8").decode("unicode_escape", errors="ignore").strip()
        if _is_visible_candidate(candidate):
            chunks.append(candidate)

    return " ".join(chunks)


def _read_site_content(site_dir: Path) -> tuple[str, str, str]:
    title = ""
    visible_chunks: list[str] = []
    raw_chunks: list[str] = []

    for file_path in _iter_site_source_files(site_dir):
        try:
            text = file_path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = file_path.read_text(encoding="utf-8", errors="ignore")
        raw_chunks.append(text)

        if file_path.suffix.lower() == ".html":
            html_title, html_visible = _extract_text_from_html(text)
            if not title and html_title:
                title = html_title
            if html_visible:
                visible_chunks.append(html_visible)
            continue

        candidate_text = _extract_candidate_text_from_source(text)
        if candidate_text:
            visible_chunks.append(candidate_text)

    return title, " ".join(visible_chunks), "\n".join(raw_chunks)


def _find_placeholder_reason(visible_text: str, raw_text: str) -> str:
    for pattern, reason in PLACEHOLDER_RULES:
        if pattern.search(visible_text) or pattern.search(raw_text):
            return reason
    return ""


def _load_meta(meta_path: Path, site_dir: Path) -> dict:
    if meta_path.exists():
        try:
            payload = json.loads(meta_path.read_text(encoding="utf-8"))
            if isinstance(payload, dict):
                return payload
        except Exception:
            pass

    parts = site_dir.relative_to(OUTPUT_DIR).parts
    payload = {
        "country": parts[0] if len(parts) > 0 else "",
        "city": parts[1] if len(parts) > 1 else "",
        "category": parts[2] if len(parts) > 2 else "",
        "shop_id": site_dir.name,
    }
    return payload


def _evaluate_site(site_dir: Path, output_dir: Path) -> SiteCheckResult:
    index_path = site_dir / "index.html"
    meta_path = site_dir / "_lead_meta.json"
    meta = _load_meta(meta_path, site_dir)
    category_dir_excel = site_dir.parent / "leads.xlsx"
    country = str(meta.get("country") or "").strip()
    city = str(meta.get("city") or "").strip()
    category = str(meta.get("category") or "").strip()
    if category_dir_excel.exists():
        excel_path = category_dir_excel
    elif country and city and category:
        excel_path = OUTPUT_DIR / country / city / category / "leads.xlsx"
    else:
        parts = site_dir.relative_to(output_dir).parts
        excel_path = output_dir / parts[0] / parts[1] / parts[2] / "leads.xlsx"
    shop_id = str(meta.get("shop_id") or site_dir.name).strip()
    title, visible_text, raw_text = _read_site_content(site_dir)

    reasons: list[str] = []
    if not title:
        reasons.append("missing title tag")
    if not _has_phone_or_email(raw_text):
        reasons.append("missing phone or email")

    visible_word_count = len(WORD_RE.findall(visible_text))
    if visible_word_count <= 200:
        reasons.append(f"visible text too short ({visible_word_count} words)")

    placeholder_reason = _find_placeholder_reason(visible_text, raw_text)
    if placeholder_reason:
        reasons.append(placeholder_reason)

    status = "good" if not reasons else "rejected"
    reason = "; ".join(reasons)
    return SiteCheckResult(
        site_dir=site_dir,
        index_path=index_path,
        meta_path=meta_path,
        excel_path=excel_path,
        shop_id=shop_id,
        meta=meta,
        status=status,
        reason=reason,
        local_file_url=index_path.resolve().as_uri(),
        title=title,
        visible_word_count=visible_word_count,
    )


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _ensure_column(ws: openpyxl.worksheet.worksheet.Worksheet, header_map: dict[str, int], header: str) -> int:
    existing = header_map.get(header.lower())
    if existing is not None:
        return existing
    col_idx = ws.max_column + 1
    ws.cell(1, col_idx).value = header
    header_map[header.lower()] = col_idx
    return col_idx


def _row_value(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int, col_idx: int | None) -> str:
    if col_idx is None:
        return ""
    return str(ws.cell(row_idx, col_idx).value or "").strip()


def _is_deployed_url(value: str) -> bool:
    text = str(value or "").strip()
    return text.startswith("http://") or text.startswith("https://")


def _update_excel_group(file_path: Path, group_results: list[SiteCheckResult], dry_run: bool) -> tuple[int, int]:
    logger = logging.getLogger("website_checker")
    if not file_path.exists():
        logger.warning("No leads.xlsx found for %s", file_path.parent)
        return 0, len(group_results)

    workbook = openpyxl.load_workbook(file_path)
    ws = workbook.active

    header_map = {
        _normalize_header(ws.cell(1, col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
        if _normalize_header(ws.cell(1, col_idx).value)
    }

    shop_id_col = _ensure_column(ws, header_map, "shop_id")
    review_status_col = _ensure_column(ws, header_map, "review_status")
    review_notes_col = _ensure_column(ws, header_map, "review_notes")
    website_url_col = _ensure_column(ws, header_map, "website_url")
    generated_website_col = header_map.get("generated_website")
    deployed_url_col = header_map.get("deployed_url")

    updated = 0
    missing = 0
    results_by_shop_id = {result.shop_id: result for result in group_results if result.shop_id}
    handled_shop_ids: set[str] = set()

    for row_idx in range(2, ws.max_row + 1):
        current_shop_id = _row_value(ws, row_idx, shop_id_col)
        if current_shop_id not in results_by_shop_id:
            continue

        result = results_by_shop_id[current_shop_id]
        deployed_url = ""
        for candidate in (
            _row_value(ws, row_idx, generated_website_col),
            _row_value(ws, row_idx, deployed_url_col),
            str(result.meta.get("deployed_url") or "").strip(),
            str(result.meta.get("generated_website") or "").strip(),
        ):
            if _is_deployed_url(candidate):
                deployed_url = candidate
                break

        effective_url = deployed_url or result.local_file_url
        result.deployed_url = deployed_url
        result.effective_url = effective_url

        ws.cell(row_idx, review_status_col).value = result.status
        ws.cell(row_idx, review_notes_col).value = result.reason
        ws.cell(row_idx, website_url_col).value = effective_url
        handled_shop_ids.add(current_shop_id)
        updated += 1

    missing = len([shop_id for shop_id in results_by_shop_id if shop_id not in handled_shop_ids])
    for shop_id in sorted(shop_id for shop_id in results_by_shop_id if shop_id not in handled_shop_ids):
        logger.warning("shop_id %s not found in %s", shop_id, file_path)

    if not dry_run:
        workbook.save(file_path)

    return updated, missing


def _write_meta(result: SiteCheckResult, dry_run: bool) -> None:
    meta = dict(result.meta)
    meta["status"] = result.status
    meta["reason"] = result.reason
    meta["local_file_url"] = result.local_file_url
    meta["review_status"] = result.status
    meta["review_notes"] = result.reason
    meta["review_updated_at"] = _utc_now_iso()
    if result.deployed_url:
        meta["deployed_url"] = result.deployed_url
        meta["generated_website"] = result.deployed_url
    result.meta = meta

    if dry_run:
        return

    result.meta_path.parent.mkdir(parents=True, exist_ok=True)
    result.meta_path.write_text(json.dumps(meta, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def main() -> None:
    _configure_logging()
    args = _build_parser().parse_args()
    logger = logging.getLogger("website_checker")

    output_dir = args.output_dir.resolve()
    if args.site_dir:
        site_dirs = [args.site_dir.resolve()]
    else:
        site_dirs = _iter_site_dirs(output_dir)

    if not site_dirs:
        logger.info("No generated site folders found under %s", output_dir)
        print("PIPELINE_STAT: sites_good=0")
        print("PIPELINE_STAT: sites_rejected=0")
        return

    results = [_evaluate_site(site_dir, output_dir) for site_dir in site_dirs]

    grouped: dict[Path, list[SiteCheckResult]] = {}
    for result in results:
        grouped.setdefault(result.excel_path, []).append(result)

    excel_updated = 0
    excel_missing = 0
    for file_path, group_results in grouped.items():
        updated, missing = _update_excel_group(file_path, group_results, dry_run=args.dry_run)
        excel_updated += updated
        excel_missing += missing

    for result in results:
        if not result.effective_url:
            result.effective_url = result.deployed_url or result.local_file_url
        _write_meta(result, dry_run=args.dry_run)
        if not _emit_pipeline_event("site", result.shop_id, result.status.upper(), result.reason):
            logger.info(
                "[%s] %s title=%s words=%d url=%s%s",
                result.status,
                result.site_dir,
                result.title or "<missing>",
                result.visible_word_count,
                result.effective_url,
                f" reason={result.reason}" if result.reason else "",
            )

    good = sum(1 for result in results if result.status == "good")
    rejected = sum(1 for result in results if result.status == "rejected")
    logger.info(
        "TOTAL: checked=%d good=%d rejected=%d excel_updated=%d excel_missing=%d",
        len(results),
        good,
        rejected,
        excel_updated,
        excel_missing,
    )
    print(f"PIPELINE_STAT: sites_good={good}")
    print(f"PIPELINE_STAT: sites_rejected={rejected}")


if __name__ == "__main__":
    main()
