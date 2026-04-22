from __future__ import annotations

import hashlib
import logging
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger("email_sender")


def _header_map(ws) -> dict[str, int]:
    """Return {header_name_lower: 1-based-col-index}."""
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(1, col).value or "").strip().lower()
        if val:
            result[val] = col
    return result


def _cell(ws, row: int, col_map: dict[str, int], key: str, default: str = "") -> str:
    col = col_map.get(key.lower())
    if col is None:
        return default
    return str(ws.cell(row, col).value or "").strip()


def read_excel_leads(file_path: Path) -> list[dict[str, Any]]:
    """
    Read a leads.xlsx file and return a list of lead dicts
    compatible with the existing agent.py lead format.
    Only returns rows where build_status == "deployed", review_status == "approved",
    and email is set.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    col_map = _header_map(ws)
    leads: list[dict[str, Any]] = []

    for row_idx in range(2, ws.max_row + 1):
        build_status = _cell(ws, row_idx, col_map, "build_status")
        if build_status.lower() != "deployed":
            continue
        review_status = _cell(ws, row_idx, col_map, "review_status")
        if review_status.lower() != "approved":
            continue
        email = _cell(ws, row_idx, col_map, "email")
        if not email:
            continue

        lead: dict[str, Any] = {
            "lead_id": _cell(ws, row_idx, col_map, "shop_id"),
            "shop_name": _cell(ws, row_idx, col_map, "name"),
            "primary_email": email,
            "emails": [email],
            "phone": _cell(ws, row_idx, col_map, "phone"),
            "category": _cell(ws, row_idx, col_map, "category"),
            "location": _cell(ws, row_idx, col_map, "city"),
            "city": _cell(ws, row_idx, col_map, "city"),
            "country": _cell(ws, row_idx, col_map, "country"),
            "website_status": "none",
            "generated_website": _cell(ws, row_idx, col_map, "generated_website"),
            "whatsapp": _cell(ws, row_idx, col_map, "whatsapp"),
            "review_status": review_status,
            "rating": _cell(ws, row_idx, col_map, "rating"),
            "review_count": None,
            "_source_file": str(file_path),
        }

        if not lead["lead_id"]:
            seed = f"{email}::{file_path}::{row_idx}"
            lead["lead_id"] = "LEAD_" + hashlib.sha1(seed.encode()).hexdigest()[:10].upper()

        leads.append(lead)

    wb.close()
    return leads


def find_output_excel_files(output_root: Path) -> list[Path]:
    """Recursively find all leads.xlsx under output_root."""
    return sorted(output_root.rglob("leads.xlsx"))


def load_all_excel_leads(output_root: Path) -> list[dict[str, Any]]:
    """Load all eligible leads from all output Excel files."""
    files = find_output_excel_files(output_root)
    all_leads: list[dict[str, Any]] = []
    for f in files:
        try:
            leads = read_excel_leads(f)
            logger.info("Loaded %d eligible leads from %s", len(leads), f)
            all_leads.extend(leads)
        except Exception as exc:  # noqa: BLE001
            logger.error("Failed to read %s: %s", f, exc)
    return all_leads

